#!/usr/bin/env python3
"""
Generate test payroll control center alert data for skill evaluation.

Creates realistic test XLSX file matching the standard alert management
export format (5-column layout). Uses human-readable validation rule names,
standard alert statuses, and realistic processor IDs.

All data is 100% synthetic — no real-world payroll data was used.

Usage:
    python generate_test_data.py --output pcc_alerts_export.xlsx

Author: CC Payroll Plugin
"""

import argparse
import random
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class AlertExportGenerator:
    """Generate realistic payroll control center alert export data.

    Matches the standard 5-column alert list export from the alert management view.
    """

    # Validation rule definitions with realistic distribution weights
    VALIDATION_RULES = [
        # Data Quality rules (weight indicates relative frequency)
        ("Employees with missing tax withholding data", 8),
        ("Employees with invalid bank account details", 5),
        ("Employees with missing time evaluation results", 6),
        ("Employees missing cost center assignment", 7),
        ("Employees with duplicate employee records", 2),

        # Compliance rules
        ("Employees with garnishment order validation errors", 4),
        ("Employees with missing state tax jurisdiction", 4),
        ("Employees with expired work authorization documents", 3),

        # Processing rules
        ("Employees exceeding overtime hours threshold", 6),
        ("Employees with retroactive change pending processing", 5),
        ("Employees with unprocessed infotype changes", 4),
        ("Employees with duplicate wage type entries", 6),

        # Financial rules
        ("Employees with negative net pay results", 4),
        ("Employees with gross pay variance exceeding threshold", 3),
        ("Employees with GL posting variance detected", 3),
        ("Employees with benefit deduction exceeding net pay", 3),
    ]

    # Standard alert statuses with distribution weights
    STATUSES = {
        "Open": 50,
        "Solution Applied": 10,
        "Resolved": 20,
        "Forwarded": 20,
    }

    # Synthetic employee first/last names
    FIRST_NAMES = [
        "John", "Mary", "Robert", "Patricia", "James", "Jennifer",
        "Michael", "Linda", "David", "Barbara", "Richard", "Jessica",
        "Joseph", "Sarah", "Thomas", "Karen", "Charles", "Nancy",
        "Christopher", "Lisa", "Daniel", "Betty", "Paul", "Sandra",
        "Mark", "Dorothy", "Donald", "Susan", "George", "Elizabeth",
        "Edward", "Sharon", "Anthony", "Deborah", "William", "Laura",
        "Ruth", "Cynthia", "Emily",
    ]

    LAST_NAMES = [
        "Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia",
        "Miller", "Davis", "Wilson", "Moore", "Taylor", "Anderson",
        "Thomas", "White", "Harris", "Martin", "Thompson", "Clark",
        "Lewis", "Allen", "Young", "King", "Wright", "Hill",
        "Flores", "Green", "Adams", "Nguyen", "Jackson", "Sanchez",
    ]

    # Processor IDs (some alerts are unassigned)
    PROCESSORS = [
        "RWILSON", "PDAVIS", "KBROWN", "LMOORE", "JSMITH",
        "PAYUSR02", "PAYUSR04", "HRUSR03", "TAXUSR01", "MJOHNSON",
    ]

    def __init__(self, seed=42):
        random.seed(seed)
        self.alerts = []

    def _personnel_number(self):
        return str(random.randint(10000000, 99999999))

    def _employee_name(self):
        return f"{random.choice(self.FIRST_NAMES)} {random.choice(self.LAST_NAMES)}"

    def _pick_status(self):
        total = sum(self.STATUSES.values())
        pick = random.randint(0, total - 1)
        cumulative = 0
        for status, weight in self.STATUSES.items():
            cumulative += weight
            if pick < cumulative:
                return status
        return "Open"

    def _pick_processor(self, status):
        """Open alerts may be unassigned; others always have a processor."""
        if status == "Open":
            return random.choice(self.PROCESSORS) if random.random() > 0.5 else None
        return random.choice(self.PROCESSORS)

    def generate(self, count=50):
        """Generate specified number of alerts."""
        rules = [r[0] for r in self.VALIDATION_RULES]
        weights = [r[1] for r in self.VALIDATION_RULES]

        for _ in range(count):
            rule = random.choices(rules, weights=weights, k=1)[0]
            status = self._pick_status()
            processor = self._pick_processor(status)

            self.alerts.append({
                "Validation Rule": rule,
                "Employee Name": self._employee_name(),
                "Personnel Number": self._personnel_number(),
                "Processor": processor,
                "Status": status,
            })

        return len(self.alerts)

    def save_to_excel(self, filepath):
        """Save alerts to XLSX with basic formatting."""
        if not HAS_OPENPYXL:
            print("ERROR: openpyxl required", file=sys.stderr)
            sys.exit(1)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Alert List"

        columns = [
            ("Validation Rule", 50),
            ("Employee Name", 22),
            ("Personnel Number", 18),
            ("Processor", 14),
            ("Status", 18),
        ]

        # Header styling
        header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        header_font = Font(name="Arial", size=10, bold=True)
        header_border = Border(bottom=Side(style="thin"), right=Side(style="thin"))
        data_font = Font(name="Arial", size=10)

        for col_idx, (col_name, col_width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_width

        # Data rows
        for row_idx, alert in enumerate(self.alerts, 2):
            for col_idx, (col_name, _) in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=alert[col_name])
                cell.font = data_font

        ws.freeze_panes = "A2"
        wb.save(filepath)
        return len(self.alerts)


def main():
    parser = argparse.ArgumentParser(description="Generate alert test data (standard 5-column format)")
    parser.add_argument("--output", default="pcc_alerts_export.xlsx")
    parser.add_argument("--count", type=int, default=50)
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()

    gen = AlertExportGenerator(seed=args.seed)
    count = gen.generate(count=args.count)
    gen.save_to_excel(args.output)

    print(f"\n=== ALERT EXPORT SUMMARY ===", file=sys.stderr)
    print(f"Total: {count} alerts", file=sys.stderr)

    from collections import Counter
    status_counts = Counter(a["Status"] for a in gen.alerts)
    print("\nBy Status:", file=sys.stderr)
    for s in ["Open", "Solution Applied", "Resolved", "Forwarded"]:
        print(f"  {s}: {status_counts.get(s, 0)}", file=sys.stderr)

    rule_counts = Counter(a["Validation Rule"] for a in gen.alerts)
    print(f"\nDistinct Validation Rules: {len(rule_counts)}", file=sys.stderr)
    print("Top rules:", file=sys.stderr)
    for rule, cnt in rule_counts.most_common(5):
        print(f"  {rule}: {cnt}", file=sys.stderr)

    print(f"\nSaved to {args.output}", file=sys.stderr)


if __name__ == "__main__":
    main()
