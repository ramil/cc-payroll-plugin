#!/usr/bin/env python3
"""
Generate comprehensive XLSX triage report from triage results JSON.

Produces multi-sheet XLSX with dashboard, prioritized alerts, root cause analysis,
processor workload, routing guide, and pattern analysis.

Usage:
    python generate_triage_report.py triage_results.json --output report.xlsx

"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter


class TriageReportGenerator:
    """Generate multi-sheet XLSX report from triage results."""

    # Color scheme
    FILL_P1 = PatternFill("solid", fgColor="FFB6C1")  # Light red
    FILL_P2 = PatternFill("solid", fgColor="FFDAB9")  # Peach
    FILL_P3 = PatternFill("solid", fgColor="FFFACD")  # Lemon
    FILL_P4 = PatternFill("solid", fgColor="E0FFE0")  # Light green
    FILL_HEADER = PatternFill("solid", fgColor="366092")
    FILL_SECTION = PatternFill("solid", fgColor="4472C4")
    FILL_BATCH = PatternFill("solid", fgColor="FFFFE0")
    FILL_BLOCKING = PatternFill("solid", fgColor="FFE0E0")

    PRIORITY_FILLS = {"P1": FILL_P1, "P2": FILL_P2, "P3": FILL_P3, "P4": FILL_P4}

    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    FONT_HEADER = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    FONT_TITLE = Font(bold=True, color="FFFFFF", size=13, name="Arial")
    FONT_BOLD = Font(bold=True, name="Arial", size=10)
    FONT_NORMAL = Font(name="Arial", size=10)
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def __init__(self, results_path: str):
        self.results = self._load_results(results_path)
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    def _load_results(self, filepath: str) -> Dict[str, Any]:
        try:
            with open(filepath, "r") as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading results file: {e}", file=sys.stderr)
            sys.exit(1)

    def _set_widths(self, ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _style_header(self, cell, fill=None):
        cell.font = self.FONT_HEADER
        cell.fill = fill or self.FILL_HEADER
        cell.alignment = self.ALIGN_CENTER
        cell.border = self.THIN_BORDER

    def _style_cell(self, cell, bold=False, fill=None):
        cell.font = self.FONT_BOLD if bold else self.FONT_NORMAL
        cell.alignment = self.ALIGN_LEFT
        cell.border = self.THIN_BORDER
        if fill:
            cell.fill = fill

    def generate_dashboard(self):
        """Sheet 1: Triage Dashboard with summary stats and key insights."""
        ws = self.wb.create_sheet("Triage Dashboard")
        s = self.results["summary"]

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"].value = "SAP PCC Payroll Alert Triage Dashboard"
        ws["A1"].font = self.FONT_TITLE
        ws["A1"].fill = self.FILL_HEADER
        ws["A1"].alignment = self.ALIGN_CENTER
        ws.row_dimensions[1].height = 30

        # Summary block
        row = 3
        ws.merge_cells(f"A{row}:B{row}")
        self._style_header(ws[f"A{row}"], self.FILL_SECTION)
        self._style_header(ws[f"B{row}"], self.FILL_SECTION)
        ws[f"A{row}"].value = "Analysis Summary"
        row += 1
        for label, val in [
            ("Total Alerts", s["total_alerts"]),
            ("Analysis Date", s["analysis_date"]),
            ("Payroll Deadline", s["payroll_deadline"]),
            ("Days to Deadline", s["days_to_deadline"]),
            ("Blocking Alerts", s["blocking_alerts"]),
            ("Open & Unassigned", s["open_unassigned"]),
        ]:
            self._style_cell(ws[f"A{row}"], bold=True)
            ws[f"A{row}"].value = label
            self._style_cell(ws[f"B{row}"])
            ws[f"B{row}"].value = val
            row += 1

        # Priority distribution
        row += 1
        ws.merge_cells(f"A{row}:B{row}")
        self._style_header(ws[f"A{row}"], self.FILL_SECTION)
        self._style_header(ws[f"B{row}"], self.FILL_SECTION)
        ws[f"A{row}"].value = "Priority Distribution"
        row += 1
        for p in ["P1", "P2", "P3", "P4"]:
            cnt = self.results["priority_counts"].get(p, 0)
            fill = self.PRIORITY_FILLS[p]
            ws[f"A{row}"].value = p
            ws[f"B{row}"].value = cnt
            self._style_cell(ws[f"A{row}"], bold=True, fill=fill)
            self._style_cell(ws[f"B{row}"], fill=fill)
            row += 1

        # Category breakdown
        row += 1
        ws.merge_cells(f"A{row}:B{row}")
        self._style_header(ws[f"A{row}"], self.FILL_SECTION)
        self._style_header(ws[f"B{row}"], self.FILL_SECTION)
        ws[f"A{row}"].value = "Category Breakdown"
        row += 1
        for cat, cnt in sorted(self.results["category_counts"].items(), key=lambda x: -x[1]):
            self._style_cell(ws[f"A{row}"], bold=True)
            ws[f"A{row}"].value = cat
            self._style_cell(ws[f"B{row}"])
            ws[f"B{row}"].value = cnt
            row += 1

        # Status breakdown
        row += 1
        ws.merge_cells(f"A{row}:B{row}")
        self._style_header(ws[f"A{row}"], self.FILL_SECTION)
        self._style_header(ws[f"B{row}"], self.FILL_SECTION)
        ws[f"A{row}"].value = "Status Breakdown"
        row += 1
        for st, cnt in sorted(self.results["status_counts"].items(), key=lambda x: -x[1]):
            self._style_cell(ws[f"A{row}"], bold=True)
            ws[f"A{row}"].value = st
            self._style_cell(ws[f"B{row}"])
            ws[f"B{row}"].value = cnt
            row += 1

        # Routing distribution
        row += 1
        ws.merge_cells(f"A{row}:B{row}")
        self._style_header(ws[f"A{row}"], self.FILL_SECTION)
        self._style_header(ws[f"B{row}"], self.FILL_SECTION)
        ws[f"A{row}"].value = "Routing Distribution"
        row += 1
        for team, cnt in sorted(self.results["team_counts"].items(), key=lambda x: -x[1]):
            self._style_cell(ws[f"A{row}"], bold=True)
            ws[f"A{row}"].value = team
            self._style_cell(ws[f"B{row}"])
            ws[f"B{row}"].value = cnt
            row += 1

        self._set_widths(ws, [32, 14, 3, 20, 20, 30])

    def generate_prioritized_alerts(self):
        """Sheet 2: All alerts sorted by priority."""
        ws = self.wb.create_sheet("Prioritized Alerts")

        ws.merge_cells("A1:J1")
        ws["A1"].value = f"Prioritized Alert List — All {self.results['summary']['total_alerts']} Alerts"
        ws["A1"].font = self.FONT_TITLE
        ws["A1"].fill = self.FILL_HEADER
        ws["A1"].alignment = self.ALIGN_CENTER
        ws.row_dimensions[1].height = 28

        headers = [
            "Priority", "Alert ID", "Validation Rule", "Category",
            "Employee Name", "Personnel #", "Processor", "Status",
            "Blocking", "Primary Team",
        ]
        for col, h in enumerate(headers, 1):
            self._style_header(ws.cell(row=2, column=col))
            ws.cell(row=2, column=col).value = h

        sorted_alerts = sorted(
            self.results["triaged"],
            key=lambda x: ({"P1": 0, "P2": 1, "P3": 2, "P4": 3}[x["priority"]], x["validation_rule"]),
        )

        for i, a in enumerate(sorted_alerts, 3):
            fill = self.PRIORITY_FILLS[a["priority"]]
            vals = [
                a["priority"], a["id"], a["validation_rule"], a["category"],
                a["employee_name"], a["personnel_number"],
                a["processor"] or "(Unassigned)", a["status"],
                "YES" if a["blocking"] else "No", a["primary_team"],
            ]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=i, column=col)
                c.value = v
                cell_fill = fill if col == 1 else (
                    self.FILL_BLOCKING if col == 9 and v == "YES" else None
                )
                self._style_cell(c, fill=cell_fill)

        self._set_widths(ws, [10, 12, 48, 18, 20, 14, 14, 18, 10, 30])

    def generate_root_cause_sheet(self):
        """Sheet 3: Root cause grouping with batch processing plan."""
        ws = self.wb.create_sheet("Root Cause & Batch Plan")

        ws.merge_cells("A1:F1")
        ws["A1"].value = "Root Cause Analysis & Batch Processing Plan"
        ws["A1"].font = self.FONT_TITLE
        ws["A1"].fill = self.FILL_HEADER
        ws["A1"].alignment = self.ALIGN_CENTER
        ws.row_dimensions[1].height = 28

        headers = [
            "Root Cause (Validation Rule)", "Count", "Batch Eligible",
            "Time Savings %", "Est. Batch Time (min)", "Alert IDs",
        ]
        for col, h in enumerate(headers, 1):
            self._style_header(ws.cell(row=2, column=col))
            ws.cell(row=2, column=col).value = h

        row = 3
        # Batch-eligible groups first
        for b in self.results["batch_opportunities"]:
            ws.cell(row=row, column=1).value = b["root_cause"]
            ws.cell(row=row, column=2).value = b["count"]
            ws.cell(row=row, column=3).value = "YES"
            ws.cell(row=row, column=4).value = f"{b['savings_pct']}%"
            ws.cell(row=row, column=5).value = b["batch_time_min"]
            ws.cell(row=row, column=6).value = ", ".join(b["alert_ids"][:6])
            for col in range(1, 7):
                self._style_cell(ws.cell(row=row, column=col), fill=self.FILL_BATCH)
            row += 1

        # Non-batch groups
        for rule, group in sorted(
            self.results["root_cause_groups"].items(), key=lambda x: -x[1]["count"]
        ):
            if not group["batch_eligible"]:
                ws.cell(row=row, column=1).value = rule
                ws.cell(row=row, column=2).value = group["count"]
                ws.cell(row=row, column=3).value = "No"
                ws.cell(row=row, column=4).value = "—"
                ws.cell(row=row, column=5).value = group["individual_time_min"]
                ws.cell(row=row, column=6).value = ""
                for col in range(1, 7):
                    self._style_cell(ws.cell(row=row, column=col))
                row += 1

        self._set_widths(ws, [50, 10, 14, 16, 20, 45])

    def generate_processor_workload(self):
        """Sheet 4: Processor workload distribution."""
        ws = self.wb.create_sheet("Processor Workload")

        ws.merge_cells("A1:G1")
        ws["A1"].value = "Processor Workload & Routing Recommendations"
        ws["A1"].font = self.FONT_TITLE
        ws["A1"].fill = self.FILL_HEADER
        ws["A1"].alignment = self.ALIGN_CENTER
        ws.row_dimensions[1].height = 28

        headers = ["Processor", "Total Alerts", "P1", "P2", "P3", "P4", "Open Alerts"]
        for col, h in enumerate(headers, 1):
            self._style_header(ws.cell(row=2, column=col))
            ws.cell(row=2, column=col).value = h

        row = 3
        for proc, load in sorted(
            self.results["processor_workload"].items(), key=lambda x: -x[1]["total"]
        ):
            ws.cell(row=row, column=1).value = proc
            ws.cell(row=row, column=2).value = load["total"]
            ws.cell(row=row, column=3).value = load["p1"]
            ws.cell(row=row, column=4).value = load["p2"]
            ws.cell(row=row, column=5).value = load["p3"]
            ws.cell(row=row, column=6).value = load["p4"]
            ws.cell(row=row, column=7).value = load["open"]

            overloaded = load["total"] >= 8
            for col in range(1, 8):
                self._style_cell(
                    ws.cell(row=row, column=col),
                    fill=self.FILL_BLOCKING if overloaded and col == 1 else None,
                    bold=(col == 1),
                )
            row += 1

        self._set_widths(ws, [18, 14, 8, 8, 8, 8, 14])

    def generate_routing_guide(self):
        """Sheet 5: Routing guide with SLA reference."""
        ws = self.wb.create_sheet("Routing Guide")

        ws.merge_cells("A1:F1")
        ws["A1"].value = "Alert Routing & Assignment Guide"
        ws["A1"].font = self.FONT_TITLE
        ws["A1"].fill = self.FILL_HEADER
        ws["A1"].alignment = self.ALIGN_CENTER
        ws.row_dimensions[1].height = 28

        headers = ["Alert ID", "Validation Rule", "Category", "Primary Team", "Complexity", "SLA Deadline"]
        for col, h in enumerate(headers, 1):
            self._style_header(ws.cell(row=2, column=col))
            ws.cell(row=2, column=col).value = h

        sorted_alerts = sorted(
            self.results["triaged"],
            key=lambda x: ({"P1": 0, "P2": 1, "P3": 2, "P4": 3}[x["priority"]], x["validation_rule"]),
        )

        row = 3
        for a in sorted_alerts:
            vals = [a["id"], a["validation_rule"], a["category"],
                    a["primary_team"], a["complexity"], a["sla_deadline"]]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col)
                c.value = v
                self._style_cell(c, fill=self.PRIORITY_FILLS[a["priority"]] if col == 1 else None)
            row += 1

        # Routing legend
        row += 2
        ws.merge_cells(f"A{row}:F{row}")
        self._style_header(ws[f"A{row}"], self.FILL_SECTION)
        for c in range(2, 7):
            self._style_header(ws.cell(row=row, column=c), self.FILL_SECTION)
        ws[f"A{row}"].value = "Routing Legend"
        row += 1
        for cat, team, desc in [
            ("Data Quality", "Data Operations Team", "HR/Payroll data specialists — PA30, PA40, PE02, PE10"),
            ("Compliance", "Compliance & Legal Team", "Tax, garnishment, regulatory — PE04, PE05, PE06"),
            ("Processing", "Payroll Operations Team", "Wage types, retro changes, schema — PT40, PT50, PT60"),
            ("Financial", "Finance & Reconciliation Team", "GL posting, accruals, net pay — FB03, FAGLL03"),
        ]:
            ws[f"A{row}"].value = cat
            ws[f"B{row}"].value = team
            ws.merge_cells(f"C{row}:F{row}")
            ws[f"C{row}"].value = desc
            self._style_cell(ws[f"A{row}"], bold=True)
            self._style_cell(ws[f"B{row}"])
            self._style_cell(ws[f"C{row}"])
            row += 1

        # SLA reference
        row += 1
        ws.merge_cells(f"A{row}:F{row}")
        self._style_header(ws[f"A{row}"], self.FILL_SECTION)
        for c in range(2, 7):
            self._style_header(ws.cell(row=row, column=c), self.FILL_SECTION)
        ws[f"A{row}"].value = "SLA Reference"
        row += 1
        for label, resp, res, esc in [
            ("P1", "15 min response", "1 hour resolution", "Escalate to on-call payroll manager immediately"),
            ("P2", "1 hour response", "4 hour resolution", "Notify team lead, assign senior specialist"),
            ("P3", "4 hour response", "1 day resolution", "Standard assignment to next available team member"),
            ("P4", "1 day response", "2 day resolution", "Standard queue management, can batch process"),
        ]:
            ws[f"A{row}"].value = label
            ws[f"B{row}"].value = resp
            ws[f"C{row}"].value = res
            ws.merge_cells(f"D{row}:F{row}")
            ws[f"D{row}"].value = esc
            self._style_cell(ws[f"A{row}"], bold=True, fill=self.PRIORITY_FILLS.get(label))
            self._style_cell(ws[f"B{row}"])
            self._style_cell(ws[f"C{row}"])
            self._style_cell(ws[f"D{row}"])
            row += 1

        self._set_widths(ws, [12, 48, 18, 30, 22, 22])

    def generate(self, output_path: str):
        """Generate complete report."""
        print("Generating Triage Dashboard...", file=sys.stderr)
        self.generate_dashboard()
        print("Generating Prioritized Alerts...", file=sys.stderr)
        self.generate_prioritized_alerts()
        print("Generating Root Cause & Batch Plan...", file=sys.stderr)
        self.generate_root_cause_sheet()
        print("Generating Processor Workload...", file=sys.stderr)
        self.generate_processor_workload()
        print("Generating Routing Guide...", file=sys.stderr)
        self.generate_routing_guide()
        self.wb.save(output_path)
        print(f"Report generated: {output_path}", file=sys.stderr)


def main():
    parser = argparse.ArgumentParser(
        description="Generate comprehensive XLSX triage report from JSON results"
    )
    parser.add_argument("input_file", help="Path to triage results JSON file")
    parser.add_argument(
        "--output",
        default="triage_report.xlsx",
        help="Output file for XLSX report",
    )

    args = parser.parse_args()

    if not Path(args.input_file).exists():
        print(f"Error: Input file not found: {args.input_file}", file=sys.stderr)
        sys.exit(1)

    generator = TriageReportGenerator(args.input_file)
    generator.generate(args.output)
    print(f"\nReport saved to {args.output}", file=sys.stderr)


if __name__ == "__main__":
    main()
