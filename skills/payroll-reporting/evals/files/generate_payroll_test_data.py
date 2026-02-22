#!/usr/bin/env python3
"""
Generate sample payroll test data for evaluation and testing.

Creates an XLSX file with ~50 employee records for January 2026 payroll.
Includes realistic salaries, overtime, bonuses, and deductions.

Usage:
    python generate_payroll_test_data.py --output payroll_data.xlsx
"""

import argparse
import random
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    exit(1)


class PayrollTestDataGenerator:
    """Generate realistic payroll test data."""

    # Employee name pool
    FIRST_NAMES = [
        'John', 'Jane', 'Michael', 'Sarah', 'David', 'Emma', 'Robert', 'Lisa',
        'James', 'Mary', 'William', 'Patricia', 'Richard', 'Jennifer', 'Joseph',
        'Linda', 'Thomas', 'Barbara', 'Charles', 'Susan', 'Christopher', 'Jessica',
        'Daniel', 'Karen', 'Matthew', 'Nancy', 'Mark', 'Betty', 'Donald', 'Margaret'
    ]

    LAST_NAMES = [
        'Smith', 'Johnson', 'Williams', 'Brown', 'Jones', 'Miller', 'Davis',
        'Rodriguez', 'Martinez', 'Hernandez', 'Lopez', 'Gonzalez', 'Wilson',
        'Anderson', 'Thomas', 'Taylor', 'Moore', 'Jackson', 'Martin', 'Lee',
        'Garcia', 'White', 'Harris', 'Sanchez', 'Clark', 'Ramirez', 'Lewis'
    ]

    # Department structure
    DEPARTMENTS = {
        '100': {'name': 'Sales', 'cost_centers': ['4100', '4101', '4102'], 'salary_range': (60000, 120000)},
        '200': {'name': 'Operations', 'cost_centers': ['4300', '4301', '4302'], 'salary_range': (45000, 90000)},
        '300': {'name': 'Finance', 'cost_centers': ['4500', '4501'], 'salary_range': (65000, 130000)},
        '100': {'name': 'Sales', 'cost_centers': ['4100', '4101', '4102'], 'salary_range': (60000, 120000)},
        '200': {'name': 'Operations', 'cost_centers': ['4300', '4301', '4302'], 'salary_range': (45000, 90000)},
    }

    def __init__(self):
        """Initialize test data generator."""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Payroll'

    def generate(self, num_employees=50, output_path='payroll_data.xlsx'):
        """Generate payroll test data."""
        print(f"Generating {num_employees} employee payroll records...")

        # Create headers
        headers = [
            'Employee ID', 'Employee Name', 'Payroll Area', 'Cost Center',
            'Department', 'Wage Type', 'Wage Type Description', 'Amount',
            'Currency', 'Pay Date', 'Status'
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set column widths
        col_widths = [12, 20, 12, 12, 15, 12, 25, 12, 10, 12, 12]
        for col_idx, width in enumerate(col_widths, 1):
            self.ws.column_dimensions[chr(64 + col_idx)].width = width

        # Generate employee data
        row = 2
        departments = ['100', '200', '300']

        for emp_num in range(1, num_employees + 1):
            emp_id = f'EMP{emp_num:05d}'
            first_name = random.choice(self.FIRST_NAMES)
            last_name = random.choice(self.LAST_NAMES)
            emp_name = f"{first_name} {last_name}"

            dept = random.choice(departments)
            cost_center = random.choice(['4100', '4101', '4102', '4200', '4300', '4301', '4400', '4500'])

            # Determine salary range by department
            if dept == '100':
                salary = random.randint(60000, 120000)
            elif dept == '200':
                salary = random.randint(45000, 90000)
            else:
                salary = random.randint(65000, 130000)

            monthly_gross = salary / 12

            # Wage types and amounts
            wage_records = [
                (1000, 'Basic Pay', monthly_gross),
            ]

            # Add overtime if in Operations (cost center 43xx)
            if cost_center.startswith('43') and random.random() > 0.7:
                overtime_hours = random.randint(5, 20)
                hourly_rate = (salary / 12) / 160
                overtime_pay = overtime_hours * hourly_rate * 1.5
                wage_records.append((1100, 'Overtime', overtime_pay))

            # Add bonus if in Sales (cost center 41xx) and random
            if cost_center.startswith('41') and random.random() > 0.8:
                bonus = random.randint(500, 3000)
                wage_records.append((1200, 'Bonus', bonus))

            # Federal income tax (approx 15% of gross)
            federal_tax = monthly_gross * 0.15
            wage_records.append((2001, 'Federal Tax', federal_tax))

            # State income tax (varies, use 4%)
            state_tax = monthly_gross * 0.04
            wage_records.append((2002, 'State Tax', state_tax))

            # FICA-OASDI (6.2%)
            fica_oasdi = monthly_gross * 0.062
            wage_records.append((2003, 'FICA-OASDI', fica_oasdi))

            # FICA-Medicare (1.45%)
            fica_medicare = monthly_gross * 0.0145
            wage_records.append((2004, 'FICA-Medicare', fica_medicare))

            # Health insurance (approx $200/month employee contribution)
            health_insurance = 200
            wage_records.append((2100, 'Health Insurance', health_insurance))

            # 401k (approx 4% employee contribution)
            k401 = monthly_gross * 0.04
            wage_records.append((2200, '401k', k401))

            # Employer FICA-OASDI (6.2%)
            er_fica_oasdi = monthly_gross * 0.062
            wage_records.append((3001, 'ER FICA-OASDI', er_fica_oasdi))

            # Employer FICA-Medicare (1.45%)
            er_fica_medicare = monthly_gross * 0.0145
            wage_records.append((3002, 'ER FICA-Medicare', er_fica_medicare))

            # Employer health insurance (approx $600/month)
            er_health_insurance = 600
            wage_records.append((3100, 'ER Health Insurance', er_health_insurance))

            # Write records
            for wage_type, description, amount in wage_records:
                self.ws.cell(row=row, column=1).value = emp_id
                self.ws.cell(row=row, column=2).value = emp_name
                self.ws.cell(row=row, column=3).value = 'P1'
                self.ws.cell(row=row, column=4).value = cost_center
                self.ws.cell(row=row, column=5).value = dept
                self.ws.cell(row=row, column=6).value = wage_type
                self.ws.cell(row=row, column=7).value = description
                self.ws.cell(row=row, column=8).value = round(amount, 2)
                self.ws.cell(row=row, column=9).value = 'USD'
                self.ws.cell(row=row, column=10).value = '2026-01-31'
                self.ws.cell(row=row, column=11).value = 'Completed'

                # Format amount as currency
                self.ws.cell(row=row, column=8).number_format = '$#,##0.00'

                row += 1

        # Freeze panes
        self.ws.freeze_panes = 'A2'

        # Save
        self.wb.save(output_path)
        print(f"Test data generated successfully: {output_path}")
        print(f"Total records: {row - 2}")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(description='Generate payroll test data')
    parser.add_argument('--output', '-o', default='payroll_data.xlsx',
                        help='Output XLSX file')
    parser.add_argument('--employees', '-e', type=int, default=50,
                        help='Number of employees to generate')

    args = parser.parse_args()

    try:
        generator = PayrollTestDataGenerator()
        generator.generate(args.employees, args.output)
    except Exception as e:
        print(f"Error: {e}")
        exit(1)


if __name__ == '__main__':
    main()
