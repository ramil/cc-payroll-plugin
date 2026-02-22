#!/usr/bin/env python3
"""
Generate test payroll data for skill evaluation.

Creates realistic payroll summary XLSX files with employee data,
wages, taxes, and benefits information for testing the payroll
reporting skill.
"""

import random
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path


class PayrollTestDataGenerator:
    """Generate realistic test payroll data."""

    DEPARTMENTS = [
        "Operations", "Customer Success", "Finance", "Human Resources", 
        "Executive", "Information Technology", "Sales"
    ]
    
    COST_CENTERS = {
        "Operations": "CC-001",
        "Customer Success": "CC-002",
        "Finance": "CC-003",
        "Human Resources": "CC-004",
        "Executive": "CC-005",
        "Information Technology": "CC-006",
        "Sales": "CC-007"
    }
    
    PAY_AREAS = ["US01", "US02", "US03"]
    
    FIRST_NAMES = [
        "John", "Jane", "Michael", "Sarah", "David", "Emily", "Robert",
        "Jessica", "William", "Amanda", "James", "Rachel", "Richard", "Lisa",
        "Joseph", "Michelle", "Thomas", "Angela", "Christopher", "Diana"
    ]
    
    LAST_NAMES = [
        "Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller",
        "Davis", "Rodriguez", "Martinez", "Hernandez", "Lopez", "Gonzalez",
        "Wilson", "Anderson", "Thomas", "Taylor", "Moore", "Jackson", "Martin"
    ]

    def __init__(self):
        self.period_date = datetime(2024, 3, 31)  # March 2024
        self.prior_period_date = datetime(2024, 2, 29)  # February 2024

    def generate_employee_data(self, num_employees=50, with_prior=True):
        """Generate employee payroll data."""
        employees = []
        
        for emp_id in range(1, num_employees + 1):
            first_name = random.choice(self.FIRST_NAMES)
            last_name = random.choice(self.LAST_NAMES)
            department = random.choice(self.DEPARTMENTS)
            
            # Assign status (mostly active, some new/terminated)
            # EE_Status: 0 = Active (includes new hires), 3 = Withdrawn/Terminated
            rand = random.random()
            if rand < 0.85:
                ee_status = 0
                status_text = "Active"
            elif rand < 0.93:
                ee_status = 0
                status_text = "New Hire"
            else:
                ee_status = 3
                status_text = "Terminated"
            
            cost_center = self.COST_CENTERS[department]
            pay_area = random.choice(self.PAY_AREAS)
            
            # Generate realistic compensation
            base_salary_ranges = {
                "Operations": (35000, 55000),
                "Customer Success": (40000, 65000),
                "Finance": (50000, 85000),
                "Human Resources": (45000, 70000),
                "Executive": (120000, 250000),
                "Information Technology": (60000, 95000),
                "Sales": (45000, 80000)
            }
            
            annual_salary = random.uniform(*base_salary_ranges[department])
            monthly_gross = annual_salary / 12
            regular_pay = round(monthly_gross, 2)
            
            # Overtime (20% chance)
            if random.random() < 0.20:
                ot_hours = random.randint(5, 20)
                hourly_rate = monthly_gross / 160
                overtime_pay = round(ot_hours * hourly_rate * 1.5, 2)
            else:
                overtime_pay = 0
            
            # Bonus (10% chance)
            bonus = round(regular_pay * random.uniform(0.05, 0.15), 2) if random.random() < 0.10 else 0
            
            total_gross = regular_pay + overtime_pay + bonus
            
            # Calculate deductions
            federal_tax = round(total_gross * 0.12, 2)  # Simplified
            state_tax = round(total_gross * 0.05, 2) if pay_area != "US03" else 0  # US03 = no state tax (like WA)
            fica_ss = round(total_gross * 0.062, 2)
            fica_medicare = round(total_gross * 0.0145, 2)
            
            # Benefits
            medical_deduction = 250 if random.random() < 0.90 else 0
            dental_deduction = 50 if random.random() < 0.75 else 0
            k401_deduction = round(total_gross * 0.06, 2) if random.random() < 0.80 else 0
            
            # Net pay
            total_deductions = (federal_tax + state_tax + fica_ss + fica_medicare + 
                              medical_deduction + dental_deduction + k401_deduction)
            net_pay = round(total_gross - total_deductions, 2)
            
            # Dates
            if status_text == "Active":
                hire_date = (self.period_date - timedelta(days=random.randint(30, 1825))).strftime("%Y-%m-%d")
                term_date = ""
            elif status_text == "New Hire":
                hire_date = (self.period_date - timedelta(days=random.randint(1, 30))).strftime("%Y-%m-%d")
                term_date = ""
            else:  # Terminated
                hire_date = (self.period_date - timedelta(days=random.randint(180, 1825))).strftime("%Y-%m-%d")
                term_date = (self.period_date - timedelta(days=random.randint(1, 10))).strftime("%Y-%m-%d")
            
            employee = {
                "Employee_ID": f"EMP{emp_id:04d}",
                "Employee_Name": f"{first_name} {last_name}",
                "Department": department,
                "Cost_Center": cost_center,
                "Pay_Area": pay_area,
                "EE_Status": ee_status,
                "Status_Text": status_text,
                "Hire_Date": hire_date,
                "Term_Date": term_date,
                "Regular_Pay": regular_pay,
                "Overtime_Pay": overtime_pay,
                "Bonus": bonus,
                "Federal_Tax": federal_tax,
                "State_Tax": state_tax,
                "FICA_SS": fica_ss,
                "FICA_Medicare": fica_medicare,
                "Medical_Deduction": medical_deduction,
                "Dental_Deduction": dental_deduction,
                "401K_Deduction": k401_deduction,
                "Net_Pay": net_pay,
                "FTE": 1.0 if random.random() < 0.95 else 0.5
            }
            
            # Add prior period data (slightly different)
            if with_prior:
                employee["Prior_Regular_Pay"] = round(regular_pay * random.uniform(0.95, 1.05), 2)
                employee["Prior_Overtime_Pay"] = round(overtime_pay * random.uniform(0.9, 1.1), 2) if overtime_pay > 0 else 0
                employee["Prior_Net_Pay"] = round(net_pay * random.uniform(0.95, 1.05), 2)
            
            employees.append(employee)
        
        return employees

    def write_to_excel(self, employees, filename, period_str="March 2024"):
        """Write employee data to Excel file."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payroll Data"
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        currency_format = '$#,##0.00'
        percent_format = '0.0%'
        
        # Header row
        headers = [
            "Employee_ID", "Employee_Name", "Department", "Cost_Center", "Pay_Area",
            "EE_Status", "Status_Text", "Hire_Date", "Term_Date", "Regular_Pay", "Overtime_Pay", "Bonus",
            "Federal_Tax", "State_Tax", "FICA_SS", "FICA_Medicare",
            "Medical_Deduction", "Dental_Deduction", "401K_Deduction", "Net_Pay", "FTE",
            "Prior_Regular_Pay", "Prior_Overtime_Pay", "Prior_Net_Pay"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Data rows
        for row_num, employee in enumerate(employees, 2):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                value = employee.get(header)
                
                # Format currency fields
                if header in ["Regular_Pay", "Overtime_Pay", "Bonus", "Federal_Tax", "State_Tax",
                             "FICA_SS", "FICA_Medicare", "Medical_Deduction", "Dental_Deduction",
                             "401K_Deduction", "Net_Pay", "Prior_Regular_Pay", "Prior_Overtime_Pay", "Prior_Net_Pay"]:
                    cell.value = value if value else 0
                    cell.number_format = currency_format
                elif header == "FTE":
                    cell.value = value if value else 1.0
                    cell.number_format = percent_format
                else:
                    cell.value = value
                
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border
        
        # Auto-adjust column widths
        for col_num, header in enumerate(headers, 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            width = max(len(header) + 2, 12)
            ws.column_dimensions[col_letter].width = width
        
        # Add summary section
        summary_row = len(employees) + 3
        ws.cell(row=summary_row, column=1).value = "PERIOD:"
        ws.cell(row=summary_row, column=2).value = period_str
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        
        ws.cell(row=summary_row + 1, column=1).value = "TOTAL_GROSS_PAY:"
        total_gross = sum([e["Regular_Pay"] + e["Overtime_Pay"] + e["Bonus"] for e in employees])
        ws.cell(row=summary_row + 1, column=2).value = total_gross
        ws.cell(row=summary_row + 1, column=2).number_format = currency_format
        
        ws.cell(row=summary_row + 2, column=1).value = "ACTIVE_HEADCOUNT:"
        active_count = len([e for e in employees if e["EE_Status"] == 0])
        ws.cell(row=summary_row + 2, column=2).value = active_count
        
        # Save file
        wb.save(filename)
        print(f"Generated {filename} with {len(employees)} employees (Period: {period_str})")


def main():
    """Generate test data files."""
    output_dir = Path(__file__).parent
    
    print("Generating payroll test data...")
    
    # Generate current period data
    generator = PayrollTestDataGenerator()
    current_employees = generator.generate_employee_data(num_employees=50, with_prior=True)
    
    current_file = output_dir / "payroll_summary.xlsx"
    generator.write_to_excel(current_employees, str(current_file), "March 2024")
    
    # Generate prior period data (subset, simulating turnover)
    prior_employees = random.sample(current_employees, min(45, len(current_employees)))
    for emp in prior_employees:
        emp["Regular_Pay"] = emp.get("Prior_Regular_Pay", emp["Regular_Pay"])
        emp["Overtime_Pay"] = emp.get("Prior_Overtime_Pay", emp["Overtime_Pay"])
        emp["Net_Pay"] = emp.get("Prior_Net_Pay", emp["Net_Pay"])
        emp["EE_Status"] = 0  # All were active in prior period
        emp["Status_Text"] = "Active"
    
    prior_file = output_dir / "payroll_prior.xlsx"
    generator.write_to_excel(prior_employees, str(prior_file), "February 2024")
    
    # Generate client-specific payroll data (simpler structure)
    client_employees = current_employees[:40]
    client_file = output_dir / "client_payroll_data.xlsx"
    
    # Add SLA tracking columns for client report
    for emp in client_employees:
        emp["Processing_Status"] = "Completed"
        emp["Data_Quality"] = "Pass"
        emp["Issue_Resolved"] = "N/A"
    
    generator.write_to_excel(client_employees, str(client_file), "March 2024 - Client XYZ")
    
    print(f"\nTest data files generated in {output_dir}:")
    print(f"  - {current_file.name}")
    print(f"  - {prior_file.name}")
    print(f"  - {client_file.name}")


if __name__ == "__main__":
    main()
