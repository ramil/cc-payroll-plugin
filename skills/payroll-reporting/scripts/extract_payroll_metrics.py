#!/usr/bin/env python3
"""
Extract payroll metrics from XLSX payroll data exports.

Reads an SAP Payroll Control Center XLSX export and calculates key payroll metrics,
aggregations, and notable items. Outputs structured JSON for report generation.

Usage:
    python extract_payroll_metrics.py payroll_data.xlsx --output metrics.json
    python extract_payroll_metrics.py payroll_data.xlsx --output metrics.json --prior prior_period.xlsx
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from collections import defaultdict, OrderedDict

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


class PayrollMetricsExtractor:
    """Extract metrics from payroll XLSX data."""

    WAGE_TYPE_CATEGORIES = {
        'earnings': (1000, 1999),
        'deductions': (2000, 2999),
        'employer_contributions': (3000, 3999),
        'informational': (4000, 9999),
    }

    def __init__(self, xlsx_path, prior_xlsx_path=None):
        """Initialize extractor with payroll data file."""
        self.xlsx_path = Path(xlsx_path)
        self.prior_xlsx_path = Path(prior_xlsx_path) if prior_xlsx_path else None

        if not self.xlsx_path.exists():
            raise FileNotFoundError(f"Payroll data file not found: {self.xlsx_path}")

        self.data = self._load_xlsx(self.xlsx_path)
        self.prior_data = self._load_xlsx(self.prior_xlsx_path) if self.prior_xlsx_path else None
        self.metrics = {}

    def _load_xlsx(self, xlsx_path):
        """Load XLSX file and extract data as list of dictionaries."""
        if not xlsx_path:
            return None

        wb = load_workbook(xlsx_path)
        ws = wb.active

        headers = []
        data = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                headers = [str(h).strip() if h else f"Column{i}" for i, h in enumerate(row, 1)]
            else:
                if any(row):  # Skip empty rows
                    record = {headers[i]: row[i] for i in range(len(headers))}
                    data.append(record)

        return data

    def _get_wage_type_category(self, wage_type):
        """Categorize wage type by 1000-4000 convention."""
        try:
            wt = int(float(str(wage_type).strip()))
        except (ValueError, AttributeError):
            return None

        for category, (min_wt, max_wt) in self.WAGE_TYPE_CATEGORIES.items():
            if min_wt <= wt <= max_wt:
                return category
        return None

    def _parse_amount(self, amount):
        """Parse amount field to float."""
        if amount is None:
            return 0.0
        try:
            return float(amount)
        except (ValueError, TypeError):
            return 0.0

    def _parse_date(self, date_str):
        """Parse date field."""
        if not date_str:
            return None
        if isinstance(date_str, str):
            try:
                return datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                return None
        return date_str

    def extract_all_metrics(self):
        """Extract all payroll metrics."""
        print(f"Processing {len(self.data)} payroll records...")

        self._calculate_totals()
        self._calculate_headcount()
        self._calculate_by_cost_center()
        self._calculate_by_department()
        self._calculate_by_wage_type()
        self._calculate_averages()
        self._calculate_tax_metrics()
        self._identify_notable_items()

        if self.prior_data:
            print(f"Processing {len(self.prior_data)} prior period records for comparison...")
            self._calculate_period_comparison()

        return self.metrics

    def _calculate_totals(self):
        """Calculate total payroll metrics."""
        totals = {
            'total_gross_pay': 0.0,
            'total_deductions': 0.0,
            'total_employer_contributions': 0.0,
            'total_employer_cost': 0.0,
            'record_count': len(self.data),
        }

        for record in self.data:
            amount = self._parse_amount(record.get('Amount', 0))
            wage_type = record.get('Wage Type', '')
            category = self._get_wage_type_category(wage_type)

            if category == 'earnings':
                totals['total_gross_pay'] += amount
            elif category == 'deductions':
                totals['total_deductions'] += amount
            elif category == 'employer_contributions':
                totals['total_employer_contributions'] += amount

        totals['total_employer_cost'] = totals['total_gross_pay'] + totals['total_employer_contributions']

        self.metrics['totals'] = totals

    def _calculate_headcount(self):
        """Calculate active headcount and workforce movements."""
        unique_employees = {}

        for record in self.data:
            emp_id = record.get('Employee ID', '')
            emp_name = record.get('Employee Name', '')

            if emp_id:
                if emp_id not in unique_employees:
                    unique_employees[emp_id] = emp_name

        active_headcount = len(unique_employees)

        headcount_metrics = {
            'active_headcount': active_headcount,
            'unique_employees': unique_employees,
            'new_hires': [],  # Would need hire date data
            'terminations': [],  # Would need termination date data
        }

        self.metrics['headcount'] = headcount_metrics

    def _calculate_by_cost_center(self):
        """Aggregate metrics by cost center."""
        by_cost_center = defaultdict(lambda: {
            'headcount': set(),
            'gross_pay': 0.0,
            'deductions': 0.0,
            'employer_contributions': 0.0,
            'total_cost': 0.0,
        })

        for record in self.data:
            cost_center = record.get('Cost Center', 'Unknown')
            emp_id = record.get('Employee ID', '')
            amount = self._parse_amount(record.get('Amount', 0))
            category = self._get_wage_type_category(record.get('Wage Type', ''))

            if emp_id:
                by_cost_center[cost_center]['headcount'].add(emp_id)

            if category == 'earnings':
                by_cost_center[cost_center]['gross_pay'] += amount
            elif category == 'deductions':
                by_cost_center[cost_center]['deductions'] += amount
            elif category == 'employer_contributions':
                by_cost_center[cost_center]['employer_contributions'] += amount

        # Convert sets to counts and calculate totals
        cc_data = {}
        for cc, data in sorted(by_cost_center.items()):
            data['headcount'] = len(data['headcount'])
            data['total_cost'] = data['gross_pay'] + data['employer_contributions']
            cc_data[str(cc)] = data

        self.metrics['by_cost_center'] = cc_data

    def _calculate_by_department(self):
        """Aggregate metrics by department."""
        by_department = defaultdict(lambda: {
            'headcount': set(),
            'gross_pay': 0.0,
            'deductions': 0.0,
            'employer_contributions': 0.0,
            'total_cost': 0.0,
        })

        for record in self.data:
            department = record.get('Department', 'Unknown')
            emp_id = record.get('Employee ID', '')
            amount = self._parse_amount(record.get('Amount', 0))
            category = self._get_wage_type_category(record.get('Wage Type', ''))

            if emp_id:
                by_department[department]['headcount'].add(emp_id)

            if category == 'earnings':
                by_department[department]['gross_pay'] += amount
            elif category == 'deductions':
                by_department[department]['deductions'] += amount
            elif category == 'employer_contributions':
                by_department[department]['employer_contributions'] += amount

        # Convert sets to counts and calculate totals
        dept_data = {}
        for dept, data in sorted(by_department.items()):
            data['headcount'] = len(data['headcount'])
            data['total_cost'] = data['gross_pay'] + data['employer_contributions']
            dept_data[str(dept)] = data

        self.metrics['by_department'] = dept_data

    def _calculate_by_wage_type(self):
        """Aggregate by wage type and category."""
        by_wage_type = defaultdict(float)
        by_category = defaultdict(float)

        for record in self.data:
            wage_type = record.get('Wage Type', '')
            wage_desc = record.get('Wage Type Description', '')
            amount = self._parse_amount(record.get('Amount', 0))

            by_wage_type[f"{wage_type}:{wage_desc}"] += amount

            category = self._get_wage_type_category(wage_type)
            if category:
                by_category[category] += amount

        self.metrics['by_wage_type'] = {k: round(v, 2) for k, v in sorted(by_wage_type.items())}
        self.metrics['by_category'] = {k: round(v, 2) for k, v in by_category.items()}

    def _calculate_averages(self):
        """Calculate per-employee averages."""
        headcount = self.metrics['headcount']['active_headcount']

        if headcount > 0:
            avg_metrics = {
                'avg_gross_pay_per_employee': round(
                    self.metrics['totals']['total_gross_pay'] / headcount, 2
                ),
                'avg_total_cost_per_employee': round(
                    self.metrics['totals']['total_employer_cost'] / headcount, 2
                ),
                'avg_deductions_per_employee': round(
                    self.metrics['totals']['total_deductions'] / headcount, 2
                ),
            }
        else:
            avg_metrics = {
                'avg_gross_pay_per_employee': 0.0,
                'avg_total_cost_per_employee': 0.0,
                'avg_deductions_per_employee': 0.0,
            }

        self.metrics['averages'] = avg_metrics

    def _calculate_tax_metrics(self):
        """Calculate tax and deduction-related metrics."""
        total_gross = self.metrics['totals']['total_gross_pay']
        total_deductions = self.metrics['totals']['total_deductions']

        if total_gross > 0:
            effective_tax_rate = round((total_deductions / total_gross) * 100, 2)
        else:
            effective_tax_rate = 0.0

        tax_metrics = {
            'total_deductions': round(total_deductions, 2),
            'effective_deduction_rate': effective_tax_rate,
        }

        self.metrics['taxes'] = tax_metrics

    def _identify_notable_items(self):
        """Identify notable patterns in the data."""
        notable = {
            'observations': [],
            'data_quality_notes': [],
        }

        # Check for overtime
        overtime_total = sum(
            self._parse_amount(r.get('Amount', 0))
            for r in self.data
            if str(r.get('Wage Type', '')).startswith('11')
        )
        if overtime_total > 0:
            notable['observations'].append(
                f"Overtime detected: ${overtime_total:,.2f} total overtime pay"
            )

        # Check for bonuses
        bonus_total = sum(
            self._parse_amount(r.get('Amount', 0))
            for r in self.data
            if str(r.get('Wage Type', '')).startswith('12')
        )
        if bonus_total > 0:
            notable['observations'].append(
                f"Bonuses detected: ${bonus_total:,.2f} total bonus payments"
            )

        # Status check
        statuses = defaultdict(int)
        for record in self.data:
            status = record.get('Status', 'Unknown')
            statuses[status] += 1

        if statuses:
            status_summary = ', '.join([f"{k}: {v}" for k, v in statuses.items()])
            notable['observations'].append(f"Processing status: {status_summary}")

        # Check for missing critical fields
        critical_fields = ['Employee ID', 'Wage Type', 'Amount', 'Pay Date']
        missing_counts = {field: 0 for field in critical_fields}

        for record in self.data:
            for field in critical_fields:
                if not record.get(field):
                    missing_counts[field] += 1

        for field, count in missing_counts.items():
            if count > 0:
                notable['data_quality_notes'].append(
                    f"Missing {field}: {count} records"
                )

        self.metrics['notable'] = notable

    def _calculate_period_comparison(self):
        """Calculate period-over-period metrics if prior data available."""
        prior_metrics = self._extract_metrics_from_data(self.prior_data)
        current_metrics = {
            'gross_pay': self.metrics['totals']['total_gross_pay'],
            'headcount': self.metrics['headcount']['active_headcount'],
            'total_cost': self.metrics['totals']['total_employer_cost'],
        }

        comparison = {}

        for metric in ['gross_pay', 'headcount', 'total_cost']:
            current = current_metrics.get(metric, 0)
            prior = prior_metrics.get(metric, 0)

            if prior > 0:
                variance_pct = round(((current - prior) / prior) * 100, 2)
            else:
                variance_pct = 0.0

            comparison[metric] = {
                'current_period': current,
                'prior_period': prior,
                'absolute_variance': round(current - prior, 2),
                'percentage_variance': variance_pct,
            }

        self.metrics['period_comparison'] = comparison

    def _extract_metrics_from_data(self, data):
        """Extract basic metrics from raw data for comparison."""
        metrics = {
            'gross_pay': 0.0,
            'headcount': len(set(r.get('Employee ID', '') for r in data if r.get('Employee ID'))),
            'total_cost': 0.0,
        }

        for record in data:
            amount = self._parse_amount(record.get('Amount', 0))
            category = self._get_wage_type_category(record.get('Wage Type', ''))

            if category == 'earnings':
                metrics['gross_pay'] += amount
            elif category == 'employer_contributions':
                metrics['total_cost'] += amount

        metrics['total_cost'] += metrics['gross_pay']
        return metrics

    def to_json(self):
        """Convert metrics to JSON-serializable format."""
        # Convert defaultdicts and sets to regular types
        json_data = json.loads(json.dumps(self.metrics, default=str))
        return json_data


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='Extract payroll metrics from XLSX payroll exports'
    )
    parser.add_argument(
        'payroll_file',
        help='Path to XLSX payroll data file'
    )
    parser.add_argument(
        '--output',
        '-o',
        default='metrics.json',
        help='Output JSON file (default: metrics.json)'
    )
    parser.add_argument(
        '--prior',
        '-p',
        help='Optional prior period XLSX file for comparison'
    )

    args = parser.parse_args()

    try:
        extractor = PayrollMetricsExtractor(args.payroll_file, args.prior)
        metrics = extractor.extract_all_metrics()

        output_path = Path(args.output)
        with open(output_path, 'w') as f:
            json.dump(metrics, f, indent=2, default=str)

        print(f"\nMetrics extracted successfully!")
        print(f"Output saved to: {output_path.absolute()}")
        print(f"\nSummary:")
        print(f"  Records processed: {metrics['totals']['record_count']}")
        print(f"  Active headcount: {metrics['headcount']['active_headcount']}")
        print(f"  Total gross pay: ${metrics['totals']['total_gross_pay']:,.2f}")
        print(f"  Total employer cost: ${metrics['totals']['total_employer_cost']:,.2f}")

    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
