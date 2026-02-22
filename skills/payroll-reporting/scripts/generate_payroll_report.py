#!/usr/bin/env python3
"""
Generate formatted XLSX payroll reports from extracted metrics JSON.

Reads JSON metrics output and generates professional XLSX reports with:
- Summary sheet with KPI dashboard
- Cost Center sheet with breakdown
- Wage Type sheet with category summary
- Employee Detail sheet with sortable data (if employee data available)
- Charts sheet with visualization data

Usage:
    python generate_payroll_report.py metrics.json --output payroll_report.xlsx --type executive
"""

import argparse
import json
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


class PayrollReportGenerator:
    """Generate formatted XLSX payroll reports."""

    # Style definitions
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SUBHEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    SUBHEADER_FONT = Font(bold=True, color="000000", size=10)
    TITLE_FONT = Font(bold=True, size=14, color="366092")
    NORMAL_FONT = Font(size=10)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    CURRENCY_FORMAT = '$#,##0.00'
    PERCENT_FORMAT = '0.00%'

    def __init__(self, metrics_json_path, report_type='executive'):
        """Initialize report generator."""
        self.metrics_json_path = Path(metrics_json_path)
        self.report_type = report_type

        if not self.metrics_json_path.exists():
            raise FileNotFoundError(f"Metrics JSON file not found: {self.metrics_json_path}")

        with open(self.metrics_json_path, 'r') as f:
            self.metrics = json.load(f)

        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet

    def generate_report(self, output_path):
        """Generate complete XLSX report."""
        print(f"Generating {self.report_type} report...")

        self._create_summary_sheet()
        self._create_cost_center_sheet()
        self._create_wage_type_sheet()
        self._create_charts_sheet()

        self.wb.save(output_path)
        print(f"Report saved to: {output_path}")

    def _create_summary_sheet(self):
        """Create summary KPI dashboard sheet."""
        ws = self.wb.create_sheet("Summary", 0)
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_margins.left = 0.75
        ws.page_margins.right = 0.75
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75

        row = 1

        # Title
        ws.merge_cells(f'A{row}:D{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = 'PAYROLL SUMMARY - KEY METRICS DASHBOARD'
        title_cell.font = self.TITLE_FONT
        ws.row_dimensions[row].height = 20
        row += 1

        # Generated date
        ws.merge_cells(f'A{row}:D{row}')
        date_cell = ws[f'A{row}']
        date_cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
        date_cell.font = Font(italic=True, size=9)
        row += 2

        # Key Metrics Table
        metrics_data = [
            ['Metric', 'Amount', 'Status'],
            ['Total Gross Pay', self.metrics['totals']['total_gross_pay'], None],
            ['Total Deductions', self.metrics['totals']['total_deductions'], None],
            ['Total Employer Cost', self.metrics['totals']['total_employer_cost'], None],
            ['Active Headcount', self.metrics['headcount']['active_headcount'], None],
            ['Average Gross Pay per Employee', self.metrics['averages']['avg_gross_pay_per_employee'], None],
            ['Average Total Cost per Employee', self.metrics['averages']['avg_total_cost_per_employee'], None],
            ['Effective Deduction Rate', self.metrics['taxes']['effective_deduction_rate'], '%'],
        ]

        for idx, row_data in enumerate(metrics_data):
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(row=row, column=col_idx + 1)
                cell.value = value
                cell.border = self.BORDER
                cell.alignment = Alignment(horizontal='left', vertical='center')

                if idx == 0:
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT
                else:
                    if col_idx == 1 and isinstance(value, (int, float)):
                        cell.number_format = self.CURRENCY_FORMAT
                    cell.font = self.NORMAL_FONT

                if col_idx == 0 and idx > 0:
                    cell.font = Font(bold=True, size=10)

            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 10

        # Add wage type category breakdown
        row += 2
        ws.merge_cells(f'A{row}:C{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = 'BREAKDOWN BY WAGE TYPE CATEGORY'
        section_cell.font = self.SUBHEADER_FONT
        section_cell.fill = self.SUBHEADER_FILL
        row += 1

        categories_data = [
            ['Category', 'Amount', ''],
            ['Earnings', self.metrics['by_category'].get('earnings', 0), None],
            ['Deductions', self.metrics['by_category'].get('deductions', 0), None],
            ['Employer Contributions', self.metrics['by_category'].get('employer_contributions', 0), None],
        ]

        for idx, row_data in enumerate(categories_data):
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(row=row, column=col_idx + 1)
                cell.value = value
                cell.border = self.BORDER

                if idx == 0:
                    cell.fill = self.SUBHEADER_FILL
                    cell.font = self.SUBHEADER_FONT
                else:
                    if col_idx == 1 and isinstance(value, (int, float)):
                        cell.number_format = self.CURRENCY_FORMAT
                    cell.font = self.NORMAL_FONT

            row += 1

    def _create_cost_center_sheet(self):
        """Create cost center breakdown sheet."""
        ws = self.wb.create_sheet("Cost Centers")

        row = 1
        ws.merge_cells(f'A{row}:F{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = 'PAYROLL BY COST CENTER'
        title_cell.font = self.TITLE_FONT
        ws.row_dimensions[row].height = 20
        row += 2

        # Headers
        headers = ['Cost Center', 'Headcount', 'Gross Pay', 'Total Cost', 'Avg Cost/Employee', 'Deductions']
        for col_idx, header in enumerate(headers):
            cell = ws.cell(row=row, column=col_idx + 1)
            cell.value = header
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

        # Data rows
        by_cost_center = self.metrics.get('by_cost_center', {})
        for cc, data in sorted(by_cost_center.items()):
            ws.cell(row=row, column=1).value = cc
            ws.cell(row=row, column=2).value = data.get('headcount', 0)
            ws.cell(row=row, column=3).value = data.get('gross_pay', 0)
            ws.cell(row=row, column=4).value = data.get('total_cost', 0)

            headcount = data.get('headcount', 1)
            avg_cost = data.get('total_cost', 0) / headcount if headcount > 0 else 0
            ws.cell(row=row, column=5).value = avg_cost

            ws.cell(row=row, column=6).value = data.get('deductions', 0)

            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = self.BORDER
                if col in [3, 4, 5, 6]:
                    cell.number_format = self.CURRENCY_FORMAT

            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 15

    def _create_wage_type_sheet(self):
        """Create wage type breakdown sheet."""
        ws = self.wb.create_sheet("Wage Types")

        row = 1
        ws.merge_cells(f'A{row}:C{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = 'PAYROLL BY WAGE TYPE'
        title_cell.font = self.TITLE_FONT
        ws.row_dimensions[row].height = 20
        row += 2

        # Headers
        headers = ['Wage Type', 'Description', 'Amount']
        for col_idx, header in enumerate(headers):
            cell = ws.cell(row=row, column=col_idx + 1)
            cell.value = header
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

        # Data rows
        by_wage_type = self.metrics.get('by_wage_type', {})
        for wt_key, amount in sorted(by_wage_type.items()):
            parts = wt_key.split(':')
            wage_type = parts[0] if parts else ''
            description = parts[1] if len(parts) > 1 else ''

            ws.cell(row=row, column=1).value = wage_type
            ws.cell(row=row, column=2).value = description
            ws.cell(row=row, column=3).value = amount

            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.border = self.BORDER
                if col == 3:
                    cell.number_format = self.CURRENCY_FORMAT

            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15

    def _create_charts_sheet(self):
        """Create charts and visualization data sheet."""
        ws = self.wb.create_sheet("Charts")

        row = 1
        ws.merge_cells(f'A{row}:C{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = 'VISUALIZATION DATA - TOP 5 COST CENTERS'
        title_cell.font = self.TITLE_FONT
        ws.row_dimensions[row].height = 20
        row += 2

        # Headers
        headers = ['Cost Center', 'Total Cost', '% of Total']
        for col_idx, header in enumerate(headers):
            cell = ws.cell(row=row, column=col_idx + 1)
            cell.value = header
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

        # Get top 5 cost centers
        by_cost_center = self.metrics.get('by_cost_center', {})
        sorted_cc = sorted(by_cost_center.items(),
                           key=lambda x: x[1].get('total_cost', 0),
                           reverse=True)[:5]

        total_cost = self.metrics['totals'].get('total_employer_cost', 1)

        for cc, data in sorted_cc:
            cost = data.get('total_cost', 0)
            pct = (cost / total_cost * 100) if total_cost > 0 else 0

            ws.cell(row=row, column=1).value = cc
            ws.cell(row=row, column=2).value = cost
            ws.cell(row=row, column=3).value = pct

            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.border = self.BORDER
                if col == 2:
                    cell.number_format = self.CURRENCY_FORMAT
                if col == 3:
                    cell.number_format = '0.0"%"'

            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='Generate formatted XLSX payroll reports from metrics JSON'
    )
    parser.add_argument(
        'metrics_file',
        help='Path to metrics JSON file'
    )
    parser.add_argument(
        '--output',
        '-o',
        default='payroll_report.xlsx',
        help='Output XLSX file (default: payroll_report.xlsx)'
    )
    parser.add_argument(
        '--type',
        '-t',
        choices=['executive', 'finance', 'operations'],
        default='executive',
        help='Report type (default: executive)'
    )

    args = parser.parse_args()

    try:
        generator = PayrollReportGenerator(args.metrics_file, args.type)
        generator.generate_report(args.output)
        print(f"Report generated successfully: {args.output}")

    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
