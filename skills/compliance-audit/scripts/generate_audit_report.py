#!/usr/bin/env python3
"""
Audit Report Generation Script

Generates comprehensive multi-sheet XLSX audit report from JSON validation results.

Usage:
    python generate_audit_report.py validation_results.json [--output audit_report.xlsx]

Author: CC Payroll Plugin
License: Proprietary
"""

import argparse
import json
import logging
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils.dataframe import dataframe_to_rows


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class AuditReportGenerator:
    """Generates comprehensive audit report from validation results."""

    def __init__(self, validation_results: Dict[str, Any]):
        """Initialize report generator with validation results."""
        self.results = validation_results
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

        # Define styles
        self.header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        self.header_font = Font(bold=True, color='FFFFFF', size=12)
        self.critical_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        self.critical_font = Font(bold=True, color='FFFFFF')
        self.high_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        self.medium_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        self.low_fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
        self.pass_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        self.pass_font = Font(bold=True, color='FFFFFF')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate(self) -> None:
        """Generate all worksheets."""
        logger.info("Generating audit report")
        self._create_executive_summary()
        self._create_critical_findings()
        self._create_all_checks_detail()
        self._create_affected_employees()
        self._create_compliance_calendar()
        self._create_sign_off_sheet()

    def save(self, output_path: str) -> None:
        """Save workbook to file."""
        self.workbook.save(output_path)
        logger.info(f"Audit report saved to {output_path}")

    def _create_executive_summary(self) -> None:
        """Create executive summary worksheet."""
        ws = self.workbook.create_sheet('Executive Summary', 0)

        # Title
        ws['A1'] = 'PAYROLL COMPLIANCE AUDIT REPORT'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        # Validation metadata
        row = 3
        ws[f'A{row}'] = 'Validation Date:'
        ws[f'B{row}'] = self.results.get('validation_date', 'N/A')
        row += 1
        ws[f'A{row}'] = 'Payroll File:'
        ws[f'B{row}'] = self.results.get('payroll_file', 'N/A')
        row += 1
        ws[f'A{row}'] = 'Total Records Validated:'
        ws[f'B{row}'] = self.results.get('total_records', 0)
        row += 2

        # Risk Assessment
        ws[f'A{row}'] = 'RISK ASSESSMENT'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        risk_score = self.results.get('risk_score', 0)
        ws[f'A{row}'] = 'Risk Score:'
        ws[f'B{row}'] = risk_score
        ws[f'B{row}'].font = Font(bold=True, size=14)

        # Color code risk score
        if risk_score <= 20:
            ws[f'B{row}'].fill = self.pass_fill
            ws[f'B{row}'].font = Font(bold=True, size=14, color='FFFFFF')
        elif risk_score <= 40:
            ws[f'B{row}'].fill = self.low_fill
        elif risk_score <= 60:
            ws[f'B{row}'].fill = self.high_fill
        elif risk_score <= 80:
            ws[f'B{row}'].fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
            ws[f'B{row}'].font = Font(bold=True, size=14, color='FFFFFF')
        else:
            ws[f'B{row}'].fill = self.critical_fill
            ws[f'B{row}'].font = Font(bold=True, size=14, color='FFFFFF')

        row += 1
        ws[f'A{row}'] = 'Risk Level:'
        ws[f'B{row}'] = self.results.get('risk_level', 'Unknown')
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = 'Overall Status:'
        ws[f'B{row}'] = self.results.get('overall_status', 'UNKNOWN')
        if self.results.get('overall_status') == 'PASS':
            ws[f'B{row}'].fill = self.pass_fill
            ws[f'B{row}'].font = self.pass_font
        else:
            ws[f'B{row}'].fill = self.critical_fill
            ws[f'B{row}'].font = Font(bold=True, color='FFFFFF')
        row += 2

        # Check Summary
        ws[f'A{row}'] = 'VALIDATION SUMMARY'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        ws[f'A{row}'] = 'Passed Checks:'
        ws[f'B{row}'] = self.results.get('pass_count', 0)
        ws[f'B{row}'].fill = self.pass_fill
        ws[f'B{row}'].font = self.pass_font
        row += 1

        ws[f'A{row}'] = 'Warnings:'
        ws[f'B{row}'] = self.results.get('warning_count', 0)
        ws[f'B{row}'].fill = self.medium_fill
        row += 1

        ws[f'A{row}'] = 'Failed Checks:'
        ws[f'B{row}'] = self.results.get('fail_count', 0)
        if self.results.get('fail_count', 0) > 0:
            ws[f'B{row}'].fill = self.critical_fill
            ws[f'B{row}'].font = Font(bold=True, color='FFFFFF')
        row += 2

        # Category Summary
        ws[f'A{row}'] = 'CATEGORY SUMMARY'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        categories = self.results.get('validation_categories', [])
        for category in categories:
            ws[f'A{row}'] = category['category']
            ws[f'B{row}'] = f"{category['passed']}/{category['total_checks']} passed"
            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50

    def _create_critical_findings(self) -> None:
        """Create critical findings worksheet."""
        ws = self.workbook.create_sheet('Critical Findings', 1)

        # Header
        ws['A1'] = 'CRITICAL FINDINGS REQUIRING IMMEDIATE ACTION'
        ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        ws['A1'].fill = self.critical_fill
        ws.merge_cells('A1:E1')

        # Column headers
        headers = ['Category', 'Check Name', 'Affected Count', 'Details', 'Recommendation']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Data
        row = 4
        findings = self.results.get('critical_findings', [])
        if findings:
            for finding in findings:
                ws.cell(row=row, column=1, value=finding.get('category', 'N/A'))
                ws.cell(row=row, column=2, value=finding.get('check_name', 'N/A'))
                ws.cell(row=row, column=3, value=finding.get('affected_count', 0))
                ws.cell(row=row, column=4, value=finding.get('details', ''))
                ws.cell(row=row, column=5, value=finding.get('recommendation', ''))

                # Format row
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = self.critical_fill
                    cell.font = Font(color='FFFFFF')
                    cell.border = self.border
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

                row += 1
        else:
            ws.cell(row=4, column=1, value='No critical findings')
            ws.cell(row=4, column=1).fill = self.pass_fill
            ws.cell(row=4, column=1).font = self.pass_font

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 40

    def _create_all_checks_detail(self) -> None:
        """Create all checks detail worksheet."""
        ws = self.workbook.create_sheet('All Checks Detail', 2)

        # Title
        ws['A1'] = 'COMPLETE VALIDATION CHECKS DETAIL'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')

        row = 3
        categories = self.results.get('validation_categories', [])

        for category in categories:
            # Category header
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws[f'A{row}']
            cell.value = category['category']
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = self.header_fill
            cell.border = self.border
            row += 1

            # Check results for this category
            checks = category.get('checks', [])
            for check in checks:
                ws.cell(row=row, column=1, value=check.get('check_name', 'N/A'))
                ws.cell(row=row, column=2, value=check.get('status', 'UNKNOWN'))
                ws.cell(row=row, column=3, value=check.get('severity', 'Unknown'))
                ws.cell(row=row, column=4, value=check.get('affected_count', 0))
                ws.cell(row=row, column=5, value=check.get('details', ''))

                # Format cells
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.border = self.border
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

                    # Color by severity
                    severity = check.get('severity', 'Low')
                    if severity == 'Critical':
                        cell.fill = self.critical_fill
                        cell.font = Font(color='FFFFFF')
                    elif severity == 'High':
                        cell.fill = self.high_fill
                    elif severity == 'Medium':
                        cell.fill = self.medium_fill
                    elif check.get('status') == 'PASS':
                        cell.fill = self.pass_fill
                        cell.font = Font(color='FFFFFF')

                row += 1

            row += 1  # Space between categories

        # Column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 50

    def _create_affected_employees(self) -> None:
        """Create affected employees worksheet."""
        ws = self.workbook.create_sheet('Affected Employees', 3)

        # Title
        ws['A1'] = 'EMPLOYEES WITH VALIDATION ISSUES'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')

        row = 3
        affected = self.results.get('affected_employees', {})

        if affected:
            for category, employees in affected.items():
                # Category header
                ws.merge_cells(f'A{row}:B{row}')
                cell = ws[f'A{row}']
                cell.value = category
                cell.font = Font(bold=True, size=11, color='FFFFFF')
                cell.fill = self.header_fill
                cell.border = self.border
                row += 1

                # Employee list
                for emp in set(employees):  # Remove duplicates
                    ws.cell(row=row, column=1, value=emp)
                    ws.cell(row=row, column=1).border = self.border
                    row += 1

                row += 1
        else:
            ws.cell(row=3, column=1, value='No employees with issues')
            ws.cell(row=3, column=1).fill = self.pass_fill
            ws.cell(row=3, column=1).font = self.pass_font

        ws.column_dimensions['A'].width = 40

    def _create_compliance_calendar(self) -> None:
        """Create compliance calendar worksheet."""
        ws = self.workbook.create_sheet('Compliance Calendar', 4)

        # Title
        ws['A1'] = 'PAYROLL COMPLIANCE CALENDAR - 2025'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')

        # Headers
        headers = ['Filing/Obligation', 'Deadline', 'Frequency', 'Penalty for Late Filing']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border

        # Compliance calendar data
        calendar_data = [
            ['Form 941 (Quarterly)', 'Month following quarter', 'Quarterly', '5-10% of unpaid taxes + interest'],
            ['Federal Payroll Tax Deposits', 'Last business day of month', 'Monthly/Semi-weekly', 'Accuracy and timeliness penalties'],
            ['State Income Tax Withholding', 'Varies by state', 'Monthly/Quarterly', 'State-specific penalties'],
            ['State Unemployment Insurance (SUI)', 'Quarterly', 'Quarterly', '10-15% of unreported wages'],
            ['W-2 Distribution', 'January 31', 'Annual', 'Not typically penalized'],
            ['Form W-3 (SSA Transmittal)', 'February 28', 'Annual', '$50-$100 per late filing'],
            ['Form 940 (FUTA)', 'January 31', 'Annual', '5-10% of unpaid tax + interest'],
            ['ACA Forms 1094-C/1095-C', 'February 28', 'Annual', '$100-$500 per employee'],
            ['State Annual Reconciliation', 'Varies by state', 'Annual', 'State-specific penalties'],
            ['Garnishment Payments', 'As ordered', 'Per order', 'Contempt of court charges'],
        ]

        row = 4
        for data_row in calendar_data:
            for col, value in enumerate(data_row, 1):
                ws.cell(row=row, column=col, value=value)
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 40

    def _create_sign_off_sheet(self) -> None:
        """Create sign-off worksheet for approvals."""
        ws = self.workbook.create_sheet('Sign-Off Sheet', 5)

        # Title
        ws['A1'] = 'PAYROLL COMPLIANCE AUDIT - SIGN-OFF SHEET'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')

        row = 3
        ws[f'A{row}'] = 'Report Date:'
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d')
        row += 1
        ws[f'A{row}'] = 'Risk Score:'
        ws[f'B{row}'] = self.results.get('risk_score', 0)
        row += 1
        ws[f'A{row}'] = 'Risk Level:'
        ws[f'B{row}'] = self.results.get('risk_level', 'Unknown')
        row += 3

        # Sign-off section
        ws[f'A{row}'] = 'APPROVAL SIGN-OFF'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 2

        # Preparer
        ws[f'A{row}'] = 'Preparer Name (Print):'
        ws[f'B{row}'] = '_' * 40
        row += 1
        ws[f'A{row}'] = 'Preparer Signature:'
        ws[f'B{row}'] = '_' * 40
        row += 1
        ws[f'A{row}'] = 'Date:'
        ws[f'B{row}'] = '_' * 40
        row += 2

        # Reviewer
        ws[f'A{row}'] = 'Reviewer Name (Print):'
        ws[f'B{row}'] = '_' * 40
        row += 1
        ws[f'A{row}'] = 'Reviewer Signature:'
        ws[f'B{row}'] = '_' * 40
        row += 1
        ws[f'A{row}'] = 'Date:'
        ws[f'B{row}'] = '_' * 40
        row += 2

        # Approver
        ws[f'A{row}'] = 'Approver Name (Print):'
        ws[f'B{row}'] = '_' * 40
        row += 1
        ws[f'A{row}'] = 'Approver Signature:'
        ws[f'B{row}'] = '_' * 40
        row += 1
        ws[f'A{row}'] = 'Date:'
        ws[f'B{row}'] = '_' * 40
        row += 3

        # Notes section
        ws[f'A{row}'] = 'Exceptions/Notes:'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws.merge_cells(f'A{row}:D{row+3}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='Generate audit report from validation results'
    )
    parser.add_argument('validation_file', help='Path to validation results JSON file')
    parser.add_argument('--output', default='audit_report.xlsx', help='Output XLSX file')

    args = parser.parse_args()

    # Read validation results
    try:
        with open(args.validation_file, 'r') as f:
            results = json.load(f)
        logger.info(f"Loaded validation results from {args.validation_file}")
    except Exception as e:
        logger.error(f"Failed to read validation file: {e}")
        sys.exit(1)

    # Generate report
    generator = AuditReportGenerator(results)
    generator.generate()
    generator.save(args.output)

    logger.info("Report generation complete")


if __name__ == '__main__':
    main()
