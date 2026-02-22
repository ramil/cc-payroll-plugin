#!/usr/bin/env python3
"""
Generate multi-sheet XLSX report from retroactive payroll impact analysis.

Creates summary, employee detail, by-type, GL impact, risk assessment, edge cases, and approval sheets.
"""

import sys
import json
import argparse
from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# Color definitions
HEADER_FILL = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
RISK_CRITICAL_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
RISK_HIGH_FILL = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
RISK_MEDIUM_FILL = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')
RISK_LOW_FILL = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def apply_header_style(ws, row_num: int, num_cols: int):
    """Apply header style to a row."""
    for col_num in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def apply_risk_color(cell, risk_level: str):
    """Apply color based on risk level."""
    if risk_level == 'Critical':
        cell.fill = RISK_CRITICAL_FILL
        cell.font = Font(bold=True, color='FFFFFF')
    elif risk_level == 'High':
        cell.fill = RISK_HIGH_FILL
        cell.font = Font(bold=True)
    elif risk_level == 'Medium':
        cell.fill = RISK_MEDIUM_FILL
    elif risk_level == 'Low':
        cell.fill = RISK_LOW_FILL


def add_summary_sheet(wb, analysis: Dict[str, Any]):
    """Add Summary sheet with high-level overview."""
    ws = wb.create_sheet('Summary', 0)

    # Title
    ws['A1'] = 'Retroactive Payroll Impact Analysis'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:B1')

    # Timestamp
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(italic=True, size=9)

    # Summary metrics
    summary = analysis['retro_summary']

    row = 4
    ws[f'A{row}'] = 'Total Employees Affected'
    ws[f'B{row}'] = summary['total_employees_affected']
    ws[f'B{row}'].font = Font(bold=True, size=12)
    row += 1

    ws[f'A{row}'] = 'Total Retro Amount'
    ws[f'B{row}'] = summary['total_retro_amount']
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'].number_format = '$#,##0.00'
    row += 2

    # By Retro Type
    ws[f'A{row}'] = 'By Retro Type'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    for retro_type, count in sorted(summary['by_retro_type'].items()):
        ws[f'A{row}'] = f'  {retro_type}'
        ws[f'B{row}'] = count
        row += 1

    row += 1

    # By Risk Level
    ws[f'A{row}'] = 'By Risk Level'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    for risk_level in ['Low', 'Medium', 'High', 'Critical']:
        count = summary['by_risk_level'].get(risk_level, 0)
        ws[f'A{row}'] = f'  {risk_level}'
        ws[f'B{row}'] = count
        apply_risk_color(ws[f'A{row}'], risk_level)
        row += 1

    # GL Impact
    gl_impact = analysis['gl_impact']
    row += 1
    ws[f'A{row}'] = 'GL Impact'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    ws[f'A{row}'] = 'Accounts Affected'
    ws[f'B{row}'] = len(gl_impact['estimated_accounts_affected'])
    row += 1

    ws[f'A{row}'] = 'GL Accounts'
    ws[f'B{row}'] = ', '.join(gl_impact['estimated_accounts_affected'])
    row += 1

    ws[f'A{row}'] = 'Total Difference Amount'
    ws[f'B{row}'] = gl_impact['total_difference_amount']
    ws[f'B{row}'].number_format = '$#,##0.00'

    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25


def add_employee_detail_sheet(wb, analysis: Dict[str, Any]):
    """Add Employee Detail sheet with per-employee breakdown."""
    ws = wb.create_sheet('Employee Detail')

    # Headers
    headers = ['Employee_ID', 'Employee_Name', 'Department', 'Cost_Center', 'Retro_Type',
               'Total_Delta', 'Risk_Level', 'Wage_Type', 'Prior_Amount', 'Current_Amount',
               'Delta', 'Edge_Cases']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

    row = 2
    for emp in analysis['affected_employees']:
        emp_id = emp['employee_id']
        emp_name = emp['employee_name']
        dept = emp['department']
        cc = emp['cost_center']
        retro_type = emp['retro_type']
        total_delta = emp['total_retro_delta']
        risk_level = emp['risk_level']
        edge_cases = ', '.join(emp['edge_cases']) if emp['edge_cases'] else ''

        # One row per wage type change
        wage_types = emp['wage_type_changes']
        if wage_types:
            for wt, change in sorted(wage_types.items()):
                ws.cell(row=row, column=1).value = emp_id
                ws.cell(row=row, column=2).value = emp_name
                ws.cell(row=row, column=3).value = dept
                ws.cell(row=row, column=4).value = cc
                ws.cell(row=row, column=5).value = retro_type
                ws.cell(row=row, column=6).value = total_delta
                ws.cell(row=row, column=6).number_format = '$#,##0.00'
                ws.cell(row=row, column=7).value = risk_level
                apply_risk_color(ws.cell(row=row, column=7), risk_level)
                ws.cell(row=row, column=8).value = wt
                ws.cell(row=row, column=9).value = change['prior']
                ws.cell(row=row, column=9).number_format = '$#,##0.00'
                ws.cell(row=row, column=10).value = change['current']
                ws.cell(row=row, column=10).number_format = '$#,##0.00'
                ws.cell(row=row, column=11).value = change['delta']
                ws.cell(row=row, column=11).number_format = '$#,##0.00'
                ws.cell(row=row, column=12).value = edge_cases

                # Apply borders
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).border = THIN_BORDER

                row += 1

    # Adjust column widths
    widths = [12, 20, 15, 12, 20, 15, 12, 10, 15, 15, 15, 30]
    for col_num, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width


def add_retro_type_sheet(wb, analysis: Dict[str, Any]):
    """Add By Retro Type sheet."""
    ws = wb.create_sheet('By Retro Type')

    retro_types = analysis['retro_summary']['by_retro_type']

    # Headers
    headers = ['Retro_Type', 'Count', 'Total_Amount', 'Avg_Amount', 'Low_Risk', 'Medium_Risk', 'High_Risk', 'Critical_Risk']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    row = 2
    for retro_type in sorted(retro_types.keys()):
        employees = [e for e in analysis['affected_employees'] if e['retro_type'] == retro_type]
        count = len(employees)
        total_amount = sum(e['total_retro_delta'] for e in employees)
        avg_amount = total_amount / count if count > 0 else 0

        # Risk breakdown
        risk_counts = {risk: sum(1 for e in employees if e['risk_level'] == risk)
                      for risk in ['Low', 'Medium', 'High', 'Critical']}

        ws.cell(row=row, column=1).value = retro_type
        ws.cell(row=row, column=2).value = count
        ws.cell(row=row, column=3).value = total_amount
        ws.cell(row=row, column=3).number_format = '$#,##0.00'
        ws.cell(row=row, column=4).value = avg_amount
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        ws.cell(row=row, column=5).value = risk_counts['Low']
        ws.cell(row=row, column=6).value = risk_counts['Medium']
        ws.cell(row=row, column=7).value = risk_counts['High']
        ws.cell(row=row, column=8).value = risk_counts['Critical']

        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER

        row += 1

    # Adjust column widths
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_num)].width = 15


def add_gl_impact_sheet(wb, analysis: Dict[str, Any]):
    """Add GL Impact sheet."""
    ws = wb.create_sheet('GL Impact')

    gl_impact = analysis['gl_impact']

    # Headers
    headers = ['GL_Account', 'Account_Description', 'Total_Amount']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    # GL mappings
    gl_desc_map = {
        '4100': 'Salary Expense',
        '4110': 'Hourly Wages',
        '4200': 'Payment Amount',
        '2100': 'Federal Tax Withheld',
        '2110': 'State Tax Withheld',
        '2120': 'Local Tax Withheld',
        '2200': 'Health Insurance Deduction',
        '2210': 'Retirement Deduction',
        '5100': 'Retroactive Adjustment',
        '5110': 'Subsequent Adjustment',
        '5120': 'Retro Change from Last',
        '5130': 'Retro Tax Adjustment',
    }

    row = 2
    for gl_account in sorted(gl_impact['estimated_accounts_affected']):
        ws.cell(row=row, column=1).value = gl_account
        ws.cell(row=row, column=2).value = gl_desc_map.get(gl_account, 'Unknown')
        # Would need detailed GL breakdown from analysis (simplified here)
        ws.cell(row=row, column=3).value = 0
        ws.cell(row=row, column=3).number_format = '$#,##0.00'

        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER

        row += 1

    # Cost Center breakdown
    row += 2
    ws[f'A{row}'] = 'By Cost Center'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    headers_cc = ['Cost_Center', 'Total_Amount']
    for col_num, header in enumerate(headers_cc, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    row += 1
    for cc, amount in sorted(gl_impact['by_cost_center'].items()):
        ws.cell(row=row, column=1).value = cc
        ws.cell(row=row, column=2).value = amount
        ws.cell(row=row, column=2).number_format = '$#,##0.00'

        for col in range(1, len(headers_cc) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER

        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15


def add_risk_assessment_sheet(wb, analysis: Dict[str, Any]):
    """Add Risk Assessment sheet."""
    ws = wb.create_sheet('Risk Assessment')

    headers = ['Employee_ID', 'Employee_Name', 'Risk_Level', 'Total_Delta', 'Retro_Type',
               'Recommended_Action']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    action_map = {
        'Low': 'Standard processing',
        'Medium': 'Review and approve',
        'High': 'Detailed review required',
        'Critical': 'Executive review required',
    }

    row = 2
    for emp in sorted(analysis['affected_employees'], key=lambda e: e['total_retro_delta'], reverse=True):
        risk_level = emp['risk_level']

        ws.cell(row=row, column=1).value = emp['employee_id']
        ws.cell(row=row, column=2).value = emp['employee_name']
        ws.cell(row=row, column=3).value = risk_level
        apply_risk_color(ws.cell(row=row, column=3), risk_level)
        ws.cell(row=row, column=4).value = emp['total_retro_delta']
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        ws.cell(row=row, column=5).value = emp['retro_type']
        ws.cell(row=row, column=6).value = action_map[risk_level]

        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER

        row += 1

    # Summary statistics
    row += 2
    ws[f'A{row}'] = 'Risk Summary'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    for risk_level in ['Low', 'Medium', 'High', 'Critical']:
        count = analysis['retro_summary']['by_risk_level'].get(risk_level, 0)
        ws[f'A{row}'] = f'{risk_level} Risk'
        ws[f'B{row}'] = count
        apply_risk_color(ws[f'A{row}'], risk_level)
        row += 1

    # Adjust column widths
    widths = [12, 20, 12, 15, 20, 25]
    for col_num, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width


def add_edge_cases_sheet(wb, analysis: Dict[str, Any]):
    """Add Edge Cases sheet."""
    ws = wb.create_sheet('Edge Cases')

    warnings = analysis['edge_case_warnings']

    if not warnings:
        ws['A1'] = 'No edge case warnings detected'
        ws['A1'].font = Font(italic=True)
        return

    headers = ['Employee_ID', 'Edge_Case_Warning', 'Risk_Level']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    row = 2
    for warning in sorted(warnings):
        # Parse warning format: "EMP_ID: Warning text"
        if ':' in warning:
            emp_id, warning_text = warning.split(':', 1)
            emp_id = emp_id.strip()
            warning_text = warning_text.strip()
        else:
            emp_id = ''
            warning_text = warning

        # Find risk level for this employee
        risk_level = 'High'
        if emp_id:
            emp_data = next((e for e in analysis['affected_employees'] if e['employee_id'] == emp_id), None)
            if emp_data:
                risk_level = emp_data['risk_level']

        ws.cell(row=row, column=1).value = emp_id
        ws.cell(row=row, column=2).value = warning_text
        ws.cell(row=row, column=3).value = risk_level
        apply_risk_color(ws.cell(row=row, column=3), risk_level)

        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER

        row += 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12


def add_approval_checklist_sheet(wb, analysis: Dict[str, Any]):
    """Add Approval Checklist sheet."""
    ws = wb.create_sheet('Approval Checklist')

    ws['A1'] = 'Retroactive Payroll Processing Approval Checklist'
    ws['A1'].font = Font(size=12, bold=True)
    ws.merge_cells('A1:C1')

    row = 3
    checklist_items = [
        ('Impact analysis reviewed and approved', ''),
        ('All affected employees identified and confirmed', ''),
        ('Retro type classifications verified', ''),
        ('Risk assessment reviewed', 'High and Critical risk items require executive sign-off'),
        ('Edge case warnings addressed', ''),
        ('GL impact estimated and reviewed', ''),
        ('Tax implications validated', ''),
        ('Simulation results accepted', ''),
        ('Master data verified for retro period', ''),
        ('Backup of prior payroll results taken', ''),
        ('Retro execution authorized', 'Only after all above items are completed'),
    ]

    headers = ['Item', 'Approved', 'Notes']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    row += 1
    for item, note in checklist_items:
        ws.cell(row=row, column=1).value = item
        ws.cell(row=row, column=2).value = '☐'  # Checkbox
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3).value = note

        for col in range(1, 4):
            ws.cell(row=row, column=col).border = THIN_BORDER

        row += 1

    # Sign-off section
    row += 2
    ws[f'A{row}'] = 'Sign-Off'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    signoff_items = [
        'Payroll Analyst',
        'Payroll Manager',
        'Finance Manager',
        'Executive Approval (if Critical risk items present)',
    ]

    for item in signoff_items:
        ws.cell(row=row, column=1).value = item
        ws.cell(row=row, column=2).value = 'Signature'
        ws.cell(row=row, column=3).value = 'Date'

        for col in range(1, 4):
            ws.cell(row=row, column=col).border = THIN_BORDER

        row += 1

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35


def generate_retro_report(analysis_json: str, output_xlsx: str):
    """
    Generate multi-sheet XLSX report from analysis JSON.

    Args:
        analysis_json: Path to JSON file from analyze_retro_impact.py
        output_xlsx: Path to output XLSX file
    """
    # Load analysis
    with open(analysis_json, 'r') as f:
        analysis = json.load(f)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Add sheets
    print("Creating report sheets...")
    add_summary_sheet(wb, analysis)
    add_employee_detail_sheet(wb, analysis)
    add_retro_type_sheet(wb, analysis)
    add_gl_impact_sheet(wb, analysis)
    add_risk_assessment_sheet(wb, analysis)
    add_edge_cases_sheet(wb, analysis)
    add_approval_checklist_sheet(wb, analysis)

    # Save workbook
    wb.save(output_xlsx)
    print(f"Report generated: {output_xlsx}")


def main():
    """Command-line interface."""
    parser = argparse.ArgumentParser(
        description='Generate XLSX report from retroactive payroll impact analysis'
    )
    parser.add_argument('analysis_json', help='JSON file from analyze_retro_impact.py')
    parser.add_argument('--output', '-o', default='retro_report.xlsx',
                       help='Output XLSX file (default: retro_report.xlsx)')

    args = parser.parse_args()

    try:
        generate_retro_report(args.analysis_json, args.output)
        print(f"\nReport complete: {Path(args.output).resolve()}")
        return 0

    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1


if __name__ == '__main__':
    sys.exit(main())
