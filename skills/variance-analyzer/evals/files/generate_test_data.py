#!/usr/bin/env python3
"""
Generate realistic SAP payroll result test data for variance analyzer testing.

Models output from SAP Wage Type Reporter (PC00_M99_CWTR) or custom ALV export.
Columns match standard SAP payroll result structure:
  - Personnel Number (PERNR): 8-digit zero-padded
  - Employee Name (ENAME)
  - Payroll Area (ABKRS): 2-character code for pay schedule periodicity
  - Payroll Period (FPPER): YYYYMM format
  - Cost Center (KOSTL)
  - Wage Type (LGART): 4-digit customer wage type code
  - Wage Type Text (LGTXT)
  - Amount (BETRG): Positive for earnings/employer costs, negative for deductions
  - Currency (WAERS)

Output files:
  payroll_current.xlsx    - Current period results (Feb 2026)
  payroll_prior.xlsx      - Prior period results (Jan 2026)
  payroll_combined.xlsx   - Combined view with both period amounts
  simulation_results.xlsx - Simulation run for pre-production validation
  prior_production.xlsx   - Prior production run for simulation comparison
"""

import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path


# ── Column Structure ─────────────────────────────────────────────────────

COLUMNS = [
    "Personnel Number",
    "Employee Name",
    "Payroll Area",
    "Payroll Period",
    "Cost Center",
    "Wage Type",
    "Wage Type Text",
    "Amount",
    "Currency",
]

COMBINED_COLUMNS = [
    "Personnel Number",
    "Employee Name",
    "Payroll Area",
    "Cost Center",
    "Wage Type",
    "Wage Type Text",
    "Prior Amount",
    "Current Amount",
    "Currency",
]


# ── Wage Types (aligned with wage-type-categories.md) ───────────────────
# Sign convention: earnings positive, employee deductions negative, employer costs positive

WAGE_TYPES = [
    # Earnings (positive)
    ("1000", "Basic Pay"),
    ("1100", "Overtime Premium 1.5x"),
    ("1200", "Bonus / Incentive Pay"),
    # Employee Taxes (negative)
    ("2001", "Federal Income Tax"),
    ("2002", "State Income Tax"),
    ("2003", "Social Security Tax - EE"),
    ("2004", "Medicare Tax - EE"),
    # Employee Benefits (negative)
    ("2100", "Medical Insurance - EE"),
    ("2101", "Dental Insurance - EE"),
    ("2200", "401(k) Deferral"),
    # Employer Costs (positive)
    ("3001", "Social Security Tax - ER"),
    ("3002", "Medicare Tax - ER"),
    ("3500", "FUTA"),
    ("3600", "SUTA"),
]


# ── Employees ────────────────────────────────────────────────────────────
# (PERNR, Name, Payroll Area, Cost Center, Role Type)
# PERNR: 8-digit zero-padded SAP Personnel Number
# Payroll Area: BW=Biweekly, SM=Semi-Monthly

EMPLOYEES_CORE = [
    ("00042868", "John Smith",       "BW", "4100", "operations"),
    ("00042901", "Jane Doe",         "BW", "4100", "operations"),
    ("00043015", "Robert Johnson",   "BW", "4200", "manufacturing"),
    ("00043022", "Alice Williams",   "BW", "4200", "manufacturing"),
    ("00043156", "Charles Brown",    "BW", "4300", "finance"),
    ("00043189", "Diana Davis",      "BW", "4300", "finance"),
    ("00043201", "Eduardo Martinez", "SM", "4400", "sales"),
    ("00043245", "Fiona Wilson",     "SM", "4400", "sales"),
    ("00043302", "George Lee",       "BW", "4500", "operations"),
    ("00043318", "Hannah Taylor",    "BW", "4500", "operations"),
    ("00043407", "Ibrahim Anderson", "BW", "4100", "operations"),
    ("00043425", "Julia Thomas",     "BW", "4100", "operations"),
    ("00043512", "Kevin Jackson",    "SM", "4200", "manufacturing"),
    ("00043538", "Laura White",      "SM", "4200", "manufacturing"),
    ("00043601", "Michael Harris",   "BW", "4300", "finance"),
    ("00043625", "Nina Clark",       "BW", "4300", "finance"),
    ("00043704", "Oscar Lewis",      "SM", "4400", "sales"),
    ("00043718", "Patricia Walker",  "SM", "4400", "sales"),
    ("00043802", "Quincy Hall",      "BW", "4500", "operations"),
    ("00043819", "Rachel Young",     "BW", "4500", "operations"),
    ("00043901", "Samuel King",      "BW", "4100", "operations"),
    ("00043925", "Teresa Wright",    "BW", "4200", "manufacturing"),
    ("00044002", "Ulysses Lopez",    "BW", "4200", "manufacturing"),
    ("00044018", "Victoria Scott",   "SM", "4300", "finance"),
    ("00044105", "William Green",    "BW", "4400", "sales"),
    ("00044122", "Xena Adams",       "BW", "4400", "sales"),
    ("00044201", "Yuri Nelson",      "SM", "4500", "operations"),
    ("00044225", "Zara Carter",      "SM", "4500", "operations"),
    ("00044302", "Adam Mitchell",    "BW", "4100", "operations"),
    ("00044318", "Beth Perez",       "BW", "4100", "operations"),
    ("00044405", "Carl Roberts",     "BW", "4200", "manufacturing"),
    ("00044422", "Dana Phillips",    "BW", "4200", "manufacturing"),
    ("00044501", "Ethan Campbell",   "SM", "4300", "finance"),
    ("00044518", "Francesca Parker", "SM", "4300", "finance"),
    ("00044605", "Gregory Evans",    "BW", "4400", "sales"),
    ("00044622", "Holly Edwards",    "BW", "4400", "sales"),
]

# New hires: appear in current period only
NEW_HIRES = [
    ("00044701", "Ian Sanchez",      "BW", "4100", "operations"),
    ("00044715", "Jasmine Jimenez",  "SM", "4500", "operations"),
]

# Terminated: appear in prior period only
TERMINATED = [
    ("00044801", "Keith Bennett",    "BW", "4200", "manufacturing"),
    ("00044815", "Lydia Murphy",     "SM", "4400", "sales"),
]

# ── Variance Scenarios ───────────────────────────────────────────────────
# These create realistic, investigatable variances between periods

# Employees who got raises (PERNR → multiplier)
RAISES = {
    "00043156": 1.05,   # Charles Brown: 5% raise (finance promotion)
    "00043302": 1.03,   # George Lee: 3% merit raise (operations)
}

# Employees with medical plan changes in current period
MEDICAL_CHANGES = {
    "00043425": {"medical": 312.00, "dental": 48.50},  # Julia Thomas: upgraded to family plan
    "00043704": {"medical": 195.00, "dental": 28.00},  # Oscar Lewis: downgraded plan
}


# ── Employee Master Data ─────────────────────────────────────────────────

def generate_master_data():
    """Generate stable employee attributes (salary, benefit elections, tax rates).
    These are fixed per employee and consistent across periods unless explicitly changed.
    """
    random.seed(100)  # Fixed seed for reproducible master data

    master = {}
    all_emps = EMPLOYEES_CORE + NEW_HIRES + TERMINATED

    for pernr, name, pay_area, cost_center, role_type in all_emps:
        # Base pay ranges by role (biweekly amounts)
        ranges = {
            "operations":    (2800, 4200),
            "manufacturing": (2600, 3800),
            "finance":       (3500, 5500),
            "sales":         (3000, 4500),
        }
        low, high = ranges[role_type]
        base = round(random.uniform(low, high), 2)

        # Adjust for semi-monthly (24 periods/year vs 26 for biweekly)
        if pay_area == "SM":
            base = round(base * 26 / 24, 2)

        master[pernr] = {
            "name": name,
            "pay_area": pay_area,
            "cost_center": cost_center,
            "role": role_type,
            "base_pay": base,
            "medical_ee": round(random.uniform(180, 350), 2),
            "dental_ee": round(random.uniform(25, 55), 2),
            "k401_pct": round(random.uniform(0.04, 0.10), 4),
            "fed_tax_rate": round(random.uniform(0.12, 0.18), 4),
            "state_tax_rate": round(random.uniform(0.03, 0.06), 4),
        }

    return master


# ── Data Generation ──────────────────────────────────────────────────────

def generate_period_data(master, employees, period, seed,
                         raises=None, med_changes=None, ot_factor=1.0):
    """Generate payroll result rows for one period.

    Args:
        master: Employee master data dict
        employees: List of (pernr, name, pay_area, cost_center, role) tuples
        period: Payroll period string (YYYYMM)
        seed: Random seed for period-specific variation (OT, etc.)
        raises: Dict of pernr → raise multiplier (e.g., 1.05 for 5%)
        med_changes: Dict of pernr → {"medical": amount, "dental": amount}
        ot_factor: Multiplier for overtime amounts (>1 = more OT)
    """
    random.seed(seed)
    raises = raises or {}
    med_changes = med_changes or {}

    data = []

    for emp_tuple in employees:
        pernr = emp_tuple[0]
        emp = master[pernr]

        # Base pay (apply raise if applicable)
        base_pay = emp["base_pay"]
        if pernr in raises:
            base_pay = round(base_pay * raises[pernr], 2)

        # Medical/dental (apply plan changes if applicable)
        medical = emp["medical_ee"]
        dental = emp["dental_ee"]
        if pernr in med_changes:
            medical = med_changes[pernr].get("medical", medical)
            dental = med_changes[pernr].get("dental", dental)

        # Overtime (period-specific random variation)
        role = emp["role"]
        if role in ("operations", "manufacturing"):
            ot = round(random.uniform(80, 700) * ot_factor, 2)
        elif role == "sales":
            ot = round(random.uniform(0, 200) * ot_factor, 2)
        else:
            ot = round(random.uniform(0, 120) * ot_factor, 2)

        gross = base_pay + ot

        # Generate row for each wage type
        for wt_code, wt_text in WAGE_TYPES:

            if wt_code == "1000":
                amount = base_pay

            elif wt_code == "1100":
                amount = ot
                if amount == 0:
                    continue  # Skip zero OT rows (SAP doesn't export zero-amount wage types)

            elif wt_code == "1200":
                continue  # No bonus this period (skip row entirely)

            elif wt_code == "2001":
                amount = -round(gross * emp["fed_tax_rate"], 2)

            elif wt_code == "2002":
                amount = -round(gross * emp["state_tax_rate"], 2)

            elif wt_code == "2003":
                amount = -round(gross * 0.062, 2)

            elif wt_code == "2004":
                amount = -round(gross * 0.0145, 2)

            elif wt_code == "2100":
                amount = -medical

            elif wt_code == "2101":
                amount = -dental

            elif wt_code == "2200":
                amount = -round(base_pay * emp["k401_pct"], 2)

            elif wt_code == "3001":
                amount = round(gross * 0.062, 2)

            elif wt_code == "3002":
                amount = round(gross * 0.0145, 2)

            elif wt_code == "3500":
                amount = round(gross * 0.006, 2)

            elif wt_code == "3600":
                amount = round(gross * 0.025, 2)

            else:
                continue

            data.append([
                pernr, emp["name"], emp["pay_area"], period,
                emp["cost_center"], wt_code, wt_text, amount, "USD"
            ])

    return data


# ── XLSX Creation ────────────────────────────────────────────────────────

def create_workbook(filepath, data, columns, sheet_name="Payroll Results"):
    """Create XLSX workbook with SAP-style formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header formatting (SAP-style blue)
    header_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    data_font = Font(name="Segoe UI", size=10)
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = data_font

            # Currency formatting for amount columns
            col_name = columns[col_idx - 1]
            if col_name in ("Amount", "Prior Amount", "Current Amount"):
                cell.number_format = '#,##0.00'

    # Auto-width columns
    for col_idx, col_name in enumerate(columns, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(len(col_name) + 4, 14)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(filepath)
    print(f"  Created: {filepath} ({len(data)} rows)")


# ── File Generators ──────────────────────────────────────────────────────

def generate_current_period(master):
    """Current period: February 2026.
    Includes core employees + new hires. Excludes terminated.
    Higher overtime (project deadline in CC 4500).
    Raises applied for promoted employees.
    Medical plan changes applied.
    """
    employees = EMPLOYEES_CORE + NEW_HIRES
    data = generate_period_data(
        master, employees,
        period="202602",
        seed=200,
        raises=RAISES,
        med_changes=MEDICAL_CHANGES,
        ot_factor=1.3,   # Higher OT this period (project deadline)
    )
    create_workbook("payroll_current.xlsx", data, COLUMNS)
    return data


def generate_prior_period(master):
    """Prior period: January 2026.
    Includes core employees + terminated. Excludes new hires.
    Normal overtime levels. No raises or plan changes.
    """
    employees = EMPLOYEES_CORE + TERMINATED
    data = generate_period_data(
        master, employees,
        period="202601",
        seed=201,
        raises=None,
        med_changes=None,
        ot_factor=0.8,   # Lower OT last period
    )
    create_workbook("payroll_prior.xlsx", data, COLUMNS)
    return data


def generate_combined(master):
    """Combined view: 10 employees with both period amounts side by side."""
    random.seed(300)
    subset = EMPLOYEES_CORE[:10]

    data = []
    for emp_tuple in subset:
        pernr = emp_tuple[0]
        emp = master[pernr]

        base_prior = emp["base_pay"]
        base_current = base_prior
        if pernr in RAISES:
            base_current = round(base_prior * RAISES[pernr], 2)

        # OT varies by period
        role = emp["role"]
        if role in ("operations", "manufacturing"):
            ot_prior = round(random.uniform(80, 500), 2)
            ot_current = round(random.uniform(150, 700), 2)
        else:
            ot_prior = round(random.uniform(0, 100), 2)
            ot_current = round(random.uniform(0, 150), 2)

        gross_prior = base_prior + ot_prior
        gross_current = base_current + ot_current

        medical = emp["medical_ee"]
        dental = emp["dental_ee"]
        med_current = medical
        den_current = dental
        if pernr in MEDICAL_CHANGES:
            med_current = MEDICAL_CHANGES[pernr].get("medical", medical)
            den_current = MEDICAL_CHANGES[pernr].get("dental", dental)

        for wt_code, wt_text in WAGE_TYPES:
            if wt_code == "1000":
                prior_amt = base_prior
                current_amt = base_current
            elif wt_code == "1100":
                prior_amt = ot_prior
                current_amt = ot_current
            elif wt_code == "1200":
                continue  # No bonus
            elif wt_code == "2001":
                prior_amt = -round(gross_prior * emp["fed_tax_rate"], 2)
                current_amt = -round(gross_current * emp["fed_tax_rate"], 2)
            elif wt_code == "2002":
                prior_amt = -round(gross_prior * emp["state_tax_rate"], 2)
                current_amt = -round(gross_current * emp["state_tax_rate"], 2)
            elif wt_code == "2003":
                prior_amt = -round(gross_prior * 0.062, 2)
                current_amt = -round(gross_current * 0.062, 2)
            elif wt_code == "2004":
                prior_amt = -round(gross_prior * 0.0145, 2)
                current_amt = -round(gross_current * 0.0145, 2)
            elif wt_code == "2100":
                prior_amt = -medical
                current_amt = -med_current
            elif wt_code == "2101":
                prior_amt = -dental
                current_amt = -den_current
            elif wt_code == "2200":
                prior_amt = -round(base_prior * emp["k401_pct"], 2)
                current_amt = -round(base_current * emp["k401_pct"], 2)
            elif wt_code == "3001":
                prior_amt = round(gross_prior * 0.062, 2)
                current_amt = round(gross_current * 0.062, 2)
            elif wt_code == "3002":
                prior_amt = round(gross_prior * 0.0145, 2)
                current_amt = round(gross_current * 0.0145, 2)
            elif wt_code == "3500":
                prior_amt = round(gross_prior * 0.006, 2)
                current_amt = round(gross_current * 0.006, 2)
            elif wt_code == "3600":
                prior_amt = round(gross_prior * 0.025, 2)
                current_amt = round(gross_current * 0.025, 2)
            else:
                continue

            data.append([
                pernr, emp["name"], emp["pay_area"], emp["cost_center"],
                wt_code, wt_text, prior_amt, current_amt, "USD"
            ])

    create_workbook("payroll_combined.xlsx", data, COMBINED_COLUMNS)


def generate_simulation(master):
    """Simulation results: small subset for pre-production validation.
    Simulated run has a systematic 2% error on Basic Pay (configuration issue).
    """
    random.seed(400)
    subset = EMPLOYEES_CORE[:5]
    sim_wt_codes = {"1000", "1100", "2001", "2003", "2004", "2100"}

    data = []
    for emp_tuple in subset:
        pernr = emp_tuple[0]
        emp = master[pernr]

        # Simulation: 2% higher basic pay (error condition)
        base_pay = round(emp["base_pay"] * 1.02, 2)
        ot = round(random.uniform(100, 500), 2)
        gross = base_pay + ot

        for wt_code, wt_text in WAGE_TYPES:
            if wt_code not in sim_wt_codes:
                continue

            if wt_code == "1000":
                amount = base_pay
            elif wt_code == "1100":
                amount = ot
            elif wt_code == "2001":
                amount = -round(gross * emp["fed_tax_rate"], 2)
            elif wt_code == "2003":
                amount = -round(gross * 0.062, 2)
            elif wt_code == "2004":
                amount = -round(gross * 0.0145, 2)
            elif wt_code == "2100":
                amount = -emp["medical_ee"]
            else:
                continue

            data.append([
                pernr, emp["name"], emp["pay_area"], "202602",
                emp["cost_center"], wt_code, wt_text, amount, "USD"
            ])

    create_workbook("simulation_results.xlsx", data, COLUMNS)


def generate_prior_production(master):
    """Prior production run: for comparison with simulation results."""
    random.seed(401)
    subset = EMPLOYEES_CORE[:5]
    prod_wt_codes = {"1000", "1100", "2001", "2003", "2004", "2100"}

    data = []
    for emp_tuple in subset:
        pernr = emp_tuple[0]
        emp = master[pernr]

        base_pay = emp["base_pay"]
        ot = round(random.uniform(50, 400), 2)
        gross = base_pay + ot

        for wt_code, wt_text in WAGE_TYPES:
            if wt_code not in prod_wt_codes:
                continue

            if wt_code == "1000":
                amount = base_pay
            elif wt_code == "1100":
                amount = ot
            elif wt_code == "2001":
                amount = -round(gross * emp["fed_tax_rate"], 2)
            elif wt_code == "2003":
                amount = -round(gross * 0.062, 2)
            elif wt_code == "2004":
                amount = -round(gross * 0.0145, 2)
            elif wt_code == "2100":
                amount = -emp["medical_ee"]
            else:
                continue

            data.append([
                pernr, emp["name"], emp["pay_area"], "202601",
                emp["cost_center"], wt_code, wt_text, amount, "USD"
            ])

    create_workbook("prior_production.xlsx", data, COLUMNS)


# ── Main ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Generating SAP payroll result test data...")
    print()

    master = generate_master_data()

    generate_current_period(master)
    generate_prior_period(master)
    generate_combined(master)
    generate_simulation(master)
    generate_prior_production(master)

    print()
    print("All test data files generated successfully.")
    print()
    print("Files:")
    print("  payroll_current.xlsx    - Current period (Feb 2026), core + new hires")
    print("  payroll_prior.xlsx      - Prior period (Jan 2026), core + terminated")
    print("  payroll_combined.xlsx   - Combined view (10 employees)")
    print("  simulation_results.xlsx - Simulation run (5 employees, 2% error)")
    print("  prior_production.xlsx   - Prior production run (5 employees)")
    print()
    print("Variance scenarios embedded:")
    print("  - 2 new hires (00044701, 00044715) - current only")
    print("  - 2 terminated (00044801, 00044815) - prior only")
    print("  - 1 raise: Charles Brown 5% (00043156, finance promotion)")
    print("  - 1 raise: George Lee 3% (00043302, operations merit)")
    print("  - 1 medical upgrade: Julia Thomas (00043425, family plan)")
    print("  - 1 medical downgrade: Oscar Lewis (00043704)")
    print("  - Higher overtime in current period (project deadline)")
    print("  - Simulation has 2% basic pay error vs production")
