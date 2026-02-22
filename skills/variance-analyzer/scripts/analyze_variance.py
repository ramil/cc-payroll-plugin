#!/usr/bin/env python3
"""
Payroll Variance Analysis Engine
Analyzes period-over-period payroll variance from XLSX exports
with anomaly detection, z-score analysis, and risk categorization
"""

import json
import sys
import argparse
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook
from statistics import mean, stdev

def normalize_column_name(col_name):
    """Normalize column name for matching (lowercase, remove spaces/underscores)"""
    return col_name.lower().replace(' ', '').replace('_', '')

def find_column_by_alias(headers, aliases):
    """Find column index by list of possible aliases"""
    normalized_headers = {normalize_column_name(h): i for i, h in enumerate(headers)}
    for alias in aliases:
        norm_alias = normalize_column_name(alias)
        if norm_alias in normalized_headers:
            return normalized_headers[norm_alias]
    return None

def read_xlsx_file(filepath):
    """Read XLSX file and return structured data"""
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    rows = []
    headers = None

    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [h for h in row]
        else:
            rows.append(row)

    return headers, rows

def parse_payroll_data(headers, rows):
    """Parse payroll data with flexible column detection"""
    emp_id_idx = find_column_by_alias(headers, ['Employee_ID', 'Emp_ID', 'ID', 'emp_id', 'Personnel Number', 'PERNR'])
    emp_name_idx = find_column_by_alias(headers, ['Employee_Name', 'Name', 'emp_name', 'Employee Name', 'ENAME'])
    pay_area_idx = find_column_by_alias(headers, ['Payroll_Area', 'Pay_Area', 'pay_area', 'Payroll Area', 'ABKRS'])
    cc_idx = find_column_by_alias(headers, ['Cost_Center', 'cost_center', 'CC', 'cc', 'Cost Center', 'KOSTL'])
    dept_idx = find_column_by_alias(headers, ['Department', 'Dept', 'dept'])
    wt_idx = find_column_by_alias(headers, ['Wage_Type', 'Type_Code', 'wage_code', 'wage_type', 'Wage Type', 'LGART'])
    wtd_idx = find_column_by_alias(headers, ['Wage_Type_Description', 'Description', 'Desc', 'wage_type_description', 'Wage Type Text', 'LGTXT'])
    amt_idx = find_column_by_alias(headers, ['Amount', 'amount', 'BETRG'])
    curr_idx = find_column_by_alias(headers, ['Currency', 'currency', 'WAERS'])

    records = []
    for row in rows:
        if not row or all(v is None for v in row):
            continue

        try:
            emp_id = str(row[emp_id_idx]) if emp_id_idx is not None else "UNKNOWN"
            emp_name = str(row[emp_name_idx]) if emp_name_idx is not None else "Unknown"
            pay_area = str(row[pay_area_idx]) if pay_area_idx is not None else "DEFAULT"
            cc = str(row[cc_idx]) if cc_idx is not None else "0000"
            dept = str(row[dept_idx]) if dept_idx is not None else "000"
            wt = str(row[wt_idx]) if wt_idx is not None else "0000"
            wtd = str(row[wtd_idx]) if wtd_idx is not None else ""
            amt = float(row[amt_idx]) if amt_idx is not None and row[amt_idx] is not None else 0.0
            currency = str(row[curr_idx]) if curr_idx is not None and row[curr_idx] is not None else "USD"

            records.append({
                'emp_id': emp_id,
                'emp_name': emp_name,
                'pay_area': pay_area,
                'cost_center': cc,
                'department': dept,
                'wage_type': wt,
                'wage_type_desc': wtd,
                'amount': amt,
                'currency': currency
            })
        except (ValueError, IndexError, TypeError):
            continue

    return records


def classify_risk(pct_var, abs_var, var_type):
    """Classify variance risk as HIGH, MEDIUM, or LOW based on methodology thresholds"""
    abs_pct = abs(pct_var)
    abs_amt = abs(abs_var)

    # New hires and terminations are always MEDIUM minimum
    if var_type in ('NEW', 'TERMINATION'):
        return 'HIGH' if abs_amt > 2000 else 'MEDIUM'

    if abs_pct > 10 and abs_amt > 2000:
        return 'HIGH'
    elif abs_pct > 5 or abs_amt > 500:
        return 'MEDIUM'
    elif abs_pct > 2 or abs_amt > 100:
        return 'LOW'
    return 'NONE'


def calculate_variances(current_records, prior_records, threshold_pct=5.0, threshold_abs=500.0, threshold_logic="AND"):
    """Calculate variances between current and prior period.

    threshold_logic controls how the two thresholds combine:
      - "AND" (default): flag only when BOTH pct and abs thresholds are exceeded.
      - "OR": flag when EITHER threshold is exceeded.
    """
    # Index prior records by emp_id + wage_type
    prior_index = {}
    for rec in prior_records:
        key = (rec['emp_id'], rec['wage_type'])
        prior_index[key] = rec

    # Index current records
    current_index = {}
    for rec in current_records:
        key = (rec['emp_id'], rec['wage_type'])
        current_index[key] = rec

    variances = []

    # Process current period records
    for curr_rec in current_records:
        key = (curr_rec['emp_id'], curr_rec['wage_type'])
        prior_rec = prior_index.get(key)

        prior_amt = prior_rec['amount'] if prior_rec else 0.0
        curr_amt = curr_rec['amount']
        abs_var = curr_amt - prior_amt

        # Calculate percentage variance
        if prior_amt == 0:
            pct_var = 100.0 if curr_amt != 0 else 0.0
            var_type = "NEW" if curr_amt > 0 else "NO_ACTIVITY"
        else:
            pct_var = (abs_var / abs(prior_amt)) * 100.0
            var_type = "CHANGE"

        # Check if variance exceeds thresholds
        pct_exceeded = abs(pct_var) > threshold_pct
        abs_exceeded = abs(abs_var) > threshold_abs
        if threshold_logic == "AND":
            flagged = pct_exceeded and abs_exceeded
        else:
            flagged = pct_exceeded or abs_exceeded

        risk = classify_risk(pct_var, abs_var, var_type)

        variance = {
            'emp_id': curr_rec['emp_id'],
            'emp_name': curr_rec['emp_name'],
            'pay_area': curr_rec['pay_area'],
            'cost_center': curr_rec['cost_center'],
            'department': curr_rec['department'],
            'wage_type': curr_rec['wage_type'],
            'wage_type_desc': curr_rec['wage_type_desc'],
            'prior_amount': round(prior_amt, 2),
            'current_amount': round(curr_amt, 2),
            'abs_variance': round(abs_var, 2),
            'pct_variance': round(pct_var, 1),
            'flagged': flagged,
            'variance_type': var_type,
            'risk_level': risk,
            'currency': curr_rec['currency']
        }
        variances.append(variance)

    # Detect terminations (in prior but not in current)
    for key, prior_rec in prior_index.items():
        if key not in current_index:
            abs_var = -prior_rec['amount']
            risk = classify_risk(-100.0, abs_var, 'TERMINATION')
            variances.append({
                'emp_id': prior_rec['emp_id'],
                'emp_name': prior_rec['emp_name'],
                'pay_area': prior_rec['pay_area'],
                'cost_center': prior_rec['cost_center'],
                'department': prior_rec['department'],
                'wage_type': prior_rec['wage_type'],
                'wage_type_desc': prior_rec['wage_type_desc'],
                'prior_amount': round(prior_rec['amount'], 2),
                'current_amount': 0.0,
                'abs_variance': round(abs_var, 2),
                'pct_variance': -100.0,
                'flagged': True,
                'variance_type': 'TERMINATION',
                'risk_level': risk,
                'currency': prior_rec['currency']
            })

    return variances


def detect_anomalies(variances, current_records, prior_records):
    """Detect all anomaly types: new hires, terminations, gross pay, wage type changes,
    cost center shifts, and z-score outliers. Returns a list of anomaly records."""

    anomalies = []

    # ── Helper indexes ──────────────────────────────────────────────
    # Employees in each period
    current_employees = set(r['emp_id'] for r in current_records)
    prior_employees = set(r['emp_id'] for r in prior_records)

    # Employee metadata (name, cost center) from each period
    curr_emp_meta = {}
    for r in current_records:
        if r['emp_id'] not in curr_emp_meta:
            curr_emp_meta[r['emp_id']] = {'emp_name': r['emp_name'], 'cost_center': r['cost_center'], 'department': r['department']}
    prior_emp_meta = {}
    for r in prior_records:
        if r['emp_id'] not in prior_emp_meta:
            prior_emp_meta[r['emp_id']] = {'emp_name': r['emp_name'], 'cost_center': r['cost_center'], 'department': r['department']}

    # Gross pay per employee per period (sum all wage type amounts)
    curr_gross = defaultdict(float)
    prior_gross = defaultdict(float)
    for r in current_records:
        curr_gross[r['emp_id']] += r['amount']
    for r in prior_records:
        prior_gross[r['emp_id']] += r['amount']

    # Wage types per employee per period
    curr_wage_types = defaultdict(set)
    prior_wage_types = defaultdict(set)
    for r in current_records:
        curr_wage_types[r['emp_id']].add(r['wage_type'])
    for r in prior_records:
        prior_wage_types[r['emp_id']].add(r['wage_type'])

    # Wage type descriptions lookup
    wt_desc = {}
    for r in current_records + prior_records:
        if r['wage_type'] not in wt_desc or not wt_desc[r['wage_type']]:
            wt_desc[r['wage_type']] = r['wage_type_desc']

    # ── 1. New Hires ───────────────────────────────────────────────
    new_hires = current_employees - prior_employees
    for emp_id in sorted(new_hires):
        meta = curr_emp_meta.get(emp_id, {})
        gross = round(curr_gross.get(emp_id, 0), 2)
        anomalies.append({
            'anomaly_type': 'NEW_HIRE',
            'risk_level': 'MEDIUM',
            'emp_id': emp_id,
            'emp_name': meta.get('emp_name', 'Unknown'),
            'cost_center': meta.get('cost_center', ''),
            'department': meta.get('department', ''),
            'detail': f"New employee in current period with gross pay ${gross:,.2f}",
            'impact': gross,
            'investigation': 'Verify hire date, benefits enrollment, W-4/I-9 on file, cost center assignment'
        })

    # ── 2. Terminations ───────────────────────────────────────────
    terminations = prior_employees - current_employees
    for emp_id in sorted(terminations):
        meta = prior_emp_meta.get(emp_id, {})
        gross = round(prior_gross.get(emp_id, 0), 2)
        anomalies.append({
            'anomaly_type': 'TERMINATION',
            'risk_level': 'MEDIUM',
            'emp_id': emp_id,
            'emp_name': meta.get('emp_name', 'Unknown'),
            'cost_center': meta.get('cost_center', ''),
            'department': meta.get('department', ''),
            'detail': f"Employee absent from current period; prior gross pay ${gross:,.2f}",
            'impact': -gross,
            'investigation': 'Verify termination date, final paycheck, COBRA processing, benefits termination, outstanding garnishments'
        })

    # ── 3. Gross Pay Anomalies (>30% change, excluding new/termed) ──
    continuing_employees = current_employees & prior_employees
    for emp_id in sorted(continuing_employees):
        cg = curr_gross[emp_id]
        pg = prior_gross[emp_id]
        if pg == 0:
            continue
        pct_change = ((cg - pg) / abs(pg)) * 100
        abs_change = cg - pg
        if abs(pct_change) > 30:
            meta = curr_emp_meta.get(emp_id, {})
            direction = "increased" if pct_change > 0 else "decreased"
            anomalies.append({
                'anomaly_type': 'GROSS_PAY_ANOMALY',
                'risk_level': 'HIGH',
                'emp_id': emp_id,
                'emp_name': meta.get('emp_name', 'Unknown'),
                'cost_center': meta.get('cost_center', ''),
                'department': meta.get('department', ''),
                'detail': f"Gross pay {direction} {abs(pct_change):.1f}% (${abs_change:+,.2f}): ${pg:,.2f} → ${cg:,.2f}",
                'impact': round(abs_change, 2),
                'investigation': 'Check for undocumented salary change, job code change, shift change, unpaid leave, overpayment correction, or data entry error'
            })

    # ── 4. Wage Type Appearance / Disappearance ────────────────────
    for emp_id in sorted(continuing_employees):
        curr_wts = curr_wage_types[emp_id]
        prior_wts = prior_wage_types[emp_id]
        meta = curr_emp_meta.get(emp_id, prior_emp_meta.get(emp_id, {}))

        # New wage types appearing
        appeared = curr_wts - prior_wts
        for wt in sorted(appeared):
            # Find the amount for this new wage type
            amt = 0
            for r in current_records:
                if r['emp_id'] == emp_id and r['wage_type'] == wt:
                    amt = r['amount']
                    break
            desc = wt_desc.get(wt, wt)
            anomalies.append({
                'anomaly_type': 'WAGE_TYPE_APPEARED',
                'risk_level': 'MEDIUM',
                'emp_id': emp_id,
                'emp_name': meta.get('emp_name', 'Unknown'),
                'cost_center': meta.get('cost_center', ''),
                'department': meta.get('department', ''),
                'detail': f"Wage type {wt} ({desc}) appeared — amount: ${amt:,.2f}",
                'impact': round(amt, 2),
                'investigation': 'Check for new benefit enrollment, new bonus/commission plan, payroll condition change, or misconfiguration'
            })

        # Wage types that disappeared
        disappeared = prior_wts - curr_wts
        for wt in sorted(disappeared):
            amt = 0
            for r in prior_records:
                if r['emp_id'] == emp_id and r['wage_type'] == wt:
                    amt = r['amount']
                    break
            desc = wt_desc.get(wt, wt)
            anomalies.append({
                'anomaly_type': 'WAGE_TYPE_DISAPPEARED',
                'risk_level': 'MEDIUM',
                'emp_id': emp_id,
                'emp_name': meta.get('emp_name', 'Unknown'),
                'cost_center': meta.get('cost_center', ''),
                'department': meta.get('department', ''),
                'detail': f"Wage type {wt} ({desc}) disappeared — prior amount: ${amt:,.2f}",
                'impact': round(-amt, 2),
                'investigation': 'Check for benefit termination, plan drop, payroll condition removed, or system configuration change'
            })

    # ── 5. Cost Center Shifts ──────────────────────────────────────
    for emp_id in sorted(continuing_employees):
        prior_cc = prior_emp_meta.get(emp_id, {}).get('cost_center', '')
        curr_cc = curr_emp_meta.get(emp_id, {}).get('cost_center', '')
        if prior_cc and curr_cc and prior_cc != curr_cc:
            meta = curr_emp_meta.get(emp_id, {})
            cg = round(curr_gross[emp_id], 2)
            anomalies.append({
                'anomaly_type': 'COST_CENTER_SHIFT',
                'risk_level': 'LOW',
                'emp_id': emp_id,
                'emp_name': meta.get('emp_name', 'Unknown'),
                'cost_center': curr_cc,
                'department': meta.get('department', ''),
                'detail': f"Cost center changed: {prior_cc} → {curr_cc} (current gross: ${cg:,.2f})",
                'impact': 0,
                'investigation': 'Verify transfer is documented in HR, cost center assignment is correct for new location/department'
            })

    # ── 6. Z-Score Statistical Outliers ────────────────────────────
    # Group variances by wage type and compute z-scores
    wt_groups = defaultdict(list)
    for v in variances:
        if v['variance_type'] == 'CHANGE' and v['prior_amount'] != 0:
            wt_groups[v['wage_type']].append(v)

    for wt, group in wt_groups.items():
        if len(group) < 5:
            # Need minimum sample size for meaningful z-scores
            continue
        pct_values = [v['pct_variance'] for v in group]
        mu = mean(pct_values)
        try:
            sigma = stdev(pct_values)
        except Exception:
            continue
        if sigma == 0:
            continue

        for v in group:
            z = (v['pct_variance'] - mu) / sigma
            if abs(z) > 2.0:
                severity = 'HIGH' if abs(z) > 3.0 else 'MEDIUM'
                desc = wt_desc.get(wt, wt)
                direction = "above" if z > 0 else "below"
                anomalies.append({
                    'anomaly_type': 'Z_SCORE_OUTLIER',
                    'risk_level': severity,
                    'emp_id': v['emp_id'],
                    'emp_name': v['emp_name'],
                    'cost_center': v['cost_center'],
                    'department': v['department'],
                    'detail': f"Wage type {wt} ({desc}): z-score {z:+.2f} ({direction} mean) — variance {v['pct_variance']:+.1f}% vs mean {mu:.1f}% (σ={sigma:.1f}%)",
                    'impact': v['abs_variance'],
                    'investigation': f"Statistical outlier for {desc}; variance is {abs(z):.1f} standard deviations from peer group mean"
                })

    return anomalies


def aggregate_by_dimension(variances, dimension):
    """Aggregate variances by a specific dimension"""
    aggregates = defaultdict(lambda: {'current': 0, 'prior': 0, 'variances': []})

    for var in variances:
        key = var[dimension]
        aggregates[key]['current'] += var['current_amount']
        aggregates[key]['prior'] += var['prior_amount']
        aggregates[key]['variances'].append(var)

    results = []
    for key, data in aggregates.items():
        curr = data['current']
        prior = data['prior']
        abs_var = curr - prior
        pct_var = ((curr - prior) / abs(prior)) * 100 if prior != 0 else (100.0 if curr > 0 else 0.0)

        flagged_items = [v for v in data['variances'] if v['flagged']]

        results.append({
            dimension: key,
            'current_total': round(curr, 2),
            'prior_total': round(prior, 2),
            'abs_variance': round(abs_var, 2),
            'pct_variance': round(pct_var, 1),
            'flagged_count': len(flagged_items),
            'total_count': len(data['variances'])
        })

    return sorted(results, key=lambda x: abs(x['abs_variance']), reverse=True)

def generate_summary(variances, anomalies, threshold_pct=5.0, threshold_abs=500.0, threshold_logic="AND"):
    """Generate summary statistics including anomaly counts"""
    total_current = sum(v['current_amount'] for v in variances)
    total_prior = sum(v['prior_amount'] for v in variances)
    total_variance = total_current - total_prior
    total_pct = ((total_variance / abs(total_prior)) * 100) if total_prior != 0 else 0.0

    flagged = [v for v in variances if v['flagged']]
    top_variances = sorted(flagged, key=lambda x: abs(x['abs_variance']), reverse=True)[:20]

    # Risk distribution
    risk_counts = defaultdict(int)
    for v in variances:
        risk_counts[v.get('risk_level', 'NONE')] += 1

    # Anomaly summary
    anomaly_counts = defaultdict(int)
    for a in anomalies:
        anomaly_counts[a['anomaly_type']] += 1

    return {
        'total_current': round(total_current, 2),
        'total_prior': round(total_prior, 2),
        'total_abs_variance': round(total_variance, 2),
        'total_pct_variance': round(total_pct, 1),
        'total_variances': len(variances),
        'flagged_variances': len(flagged),
        'top_variances': top_variances,
        'risk_distribution': dict(risk_counts),
        'anomaly_counts': dict(anomaly_counts),
        'total_anomalies': len(anomalies),
        'thresholds': {
            'pct': threshold_pct,
            'abs': threshold_abs,
            'logic': threshold_logic
        }
    }

def main():
    parser = argparse.ArgumentParser(
        description='Analyze payroll variance between two periods'
    )
    parser.add_argument('current_file', nargs='?', help='Current period XLSX file')
    parser.add_argument('prior_file', nargs='?', help='Prior period XLSX file')
    parser.add_argument('--threshold-pct', type=float, default=5.0,
                        help='Percentage threshold for variance flagging (default: 5)')
    parser.add_argument('--threshold-abs', type=float, default=500.0,
                        help='Absolute amount threshold for variance flagging (default: 500)')
    parser.add_argument('--threshold-logic', type=str, default='AND', choices=['AND', 'OR'],
                        help='How to combine thresholds: AND (both must exceed) or OR (either exceeds). Default: AND')
    parser.add_argument('--output', type=str, default='variance_results.json',
                        help='Output JSON file path')

    args = parser.parse_args()

    if not args.current_file or not args.prior_file:
        parser.print_help()
        sys.exit(1)

    # Load data
    print(f"Loading current period: {args.current_file}")
    curr_headers, curr_rows = read_xlsx_file(args.current_file)
    curr_records = parse_payroll_data(curr_headers, curr_rows)
    print(f"  Loaded {len(curr_records)} records")

    print(f"Loading prior period: {args.prior_file}")
    prior_headers, prior_rows = read_xlsx_file(args.prior_file)
    prior_records = parse_payroll_data(prior_headers, prior_rows)
    print(f"  Loaded {len(prior_records)} records")

    # Calculate variances
    print(f"Calculating variances (thresholds: {args.threshold_pct}% {args.threshold_logic} ${args.threshold_abs:,.0f})...")
    variances = calculate_variances(
        curr_records, prior_records,
        threshold_pct=args.threshold_pct,
        threshold_abs=args.threshold_abs,
        threshold_logic=args.threshold_logic
    )
    print(f"  Calculated {len(variances)} variance records")

    # Detect anomalies
    print("Running anomaly detection...")
    anomalies = detect_anomalies(variances, curr_records, prior_records)
    anomaly_types = defaultdict(int)
    for a in anomalies:
        anomaly_types[a['anomaly_type']] += 1
    for atype, count in sorted(anomaly_types.items()):
        print(f"  {atype}: {count}")
    print(f"  Total anomalies: {len(anomalies)}")

    # Generate aggregations
    print("Generating aggregations...")
    by_wage_type = aggregate_by_dimension(variances, 'wage_type')
    by_cost_center = aggregate_by_dimension(variances, 'cost_center')
    by_department = aggregate_by_dimension(variances, 'department')

    # Generate summary
    summary = generate_summary(variances, anomalies, args.threshold_pct, args.threshold_abs, args.threshold_logic)

    # Build output
    output = {
        'metadata': {
            'current_file': str(args.current_file),
            'prior_file': str(args.prior_file),
            'threshold_pct': args.threshold_pct,
            'threshold_abs': args.threshold_abs,
            'threshold_logic': args.threshold_logic
        },
        'summary': summary,
        'anomalies': anomalies,
        'by_wage_type': by_wage_type,
        'by_cost_center': by_cost_center,
        'by_department': by_department,
        'all_variances': variances
    }

    # Write output
    print(f"Writing results to {args.output}")
    with open(args.output, 'w') as f:
        json.dump(output, f, indent=2)

    print("\nAnalysis complete!")
    print(f"  Total payroll variance: ${summary['total_abs_variance']:,.2f} ({summary['total_pct_variance']:.1f}%)")
    print(f"  Flagged variances: {summary['flagged_variances']} of {summary['total_variances']}")
    print(f"  Risk distribution: HIGH={summary['risk_distribution'].get('HIGH', 0)}, MEDIUM={summary['risk_distribution'].get('MEDIUM', 0)}, LOW={summary['risk_distribution'].get('LOW', 0)}")
    print(f"  Anomalies detected: {summary['total_anomalies']}")

if __name__ == '__main__':
    main()
