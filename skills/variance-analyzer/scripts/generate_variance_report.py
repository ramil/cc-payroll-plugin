#!/usr/bin/env python3
"""
Variance Report Generator
Creates formatted XLSX workbook from variance analysis JSON
Includes Summary, By Wage Type, By Cost Center, By Department, Detail, and Anomalies sheets
"""

import json
import sys
import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ── Style constants ─────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

FLAG_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
HIGH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Red tint
MEDIUM_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")   # Amber tint
LOW_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")      # Green tint

RISK_FILLS = {
    'HIGH': HIGH_FILL,
    'MEDIUM': MEDIUM_FILL,
    'LOW': LOW_FILL,
}

# Anomaly type display names
ANOMALY_LABELS = {
    'NEW_HIRE': 'New Hire',
    'TERMINATION': 'Termination',
    'GROSS_PAY_ANOMALY': 'Gross Pay Anomaly',
    'WAGE_TYPE_APPEARED': 'Wage Type Appeared',
    'WAGE_TYPE_DISAPPEARED': 'Wage Type Disappeared',
    'COST_CENTER_SHIFT': 'Cost Center Shift',
    'Z_SCORE_OUTLIER': 'Statistical Outlier',
}


def format_currency(value):
    """Format as currency"""
    return f"${value:,.2f}" if isinstance(value, (int, float)) else str(value)


def add_header_row(ws, headers, row=None):
    """Add header row with formatting"""
    if row is None:
        row = ws.max_row + 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    return row


def create_summary_sheet(output, wb):
    """Create Executive Summary sheet with risk distribution and anomaly counts"""
    ws = wb.create_sheet("Summary", 0)

    summary = output['summary']

    # Title
    ws['A1'] = "PAYROLL VARIANCE ANALYSIS - SUMMARY"
    ws['A1'].font = Font(bold=True, size=14)

    # ── Key Metrics ─────────────────────────────────────────────
    row = 3
    ws[f'A{row}'] = "Total Payroll (Prior)"
    ws[f'B{row}'] = summary['total_prior']
    ws[f'B{row}'].number_format = '$#,##0.00'

    row += 1
    ws[f'A{row}'] = "Total Payroll (Current)"
    ws[f'B{row}'] = summary['total_current']
    ws[f'B{row}'].number_format = '$#,##0.00'

    row += 1
    ws[f'A{row}'] = "Absolute Variance"
    ws[f'B{row}'] = summary['total_abs_variance']
    ws[f'B{row}'].number_format = '$#,##0.00'

    row += 1
    ws[f'A{row}'] = "Percentage Variance"
    ws[f'B{row}'] = summary['total_pct_variance']
    ws[f'B{row}'].number_format = '0.0"%"'

    row += 2
    ws[f'A{row}'] = f"Flagged Variances: {summary['flagged_variances']} of {summary['total_variances']}"
    ws[f'A{row}'].font = Font(bold=True)

    # ── Risk Distribution ───────────────────────────────────────
    risk_dist = summary.get('risk_distribution', {})
    row += 2
    ws[f'A{row}'] = "RISK DISTRIBUTION"
    ws[f'A{row}'].font = Font(bold=True, size=12)

    row += 1
    for risk_level in ['HIGH', 'MEDIUM', 'LOW', 'NONE']:
        count = risk_dist.get(risk_level, 0)
        if count > 0:
            ws[f'A{row}'] = risk_level
            ws[f'B{row}'] = count
            if risk_level in RISK_FILLS:
                ws[f'A{row}'].fill = RISK_FILLS[risk_level]
            row += 1

    # ── Anomaly Summary ─────────────────────────────────────────
    anomaly_counts = summary.get('anomaly_counts', {})
    total_anomalies = summary.get('total_anomalies', 0)
    if total_anomalies > 0:
        row += 1
        ws[f'A{row}'] = f"ANOMALIES DETECTED: {total_anomalies}"
        ws[f'A{row}'].font = Font(bold=True, size=12)

        row += 1
        for atype, count in sorted(anomaly_counts.items()):
            label = ANOMALY_LABELS.get(atype, atype)
            ws[f'A{row}'] = label
            ws[f'B{row}'] = count
            row += 1

    # ── Top 20 Variances ────────────────────────────────────────
    row += 1
    ws[f'A{row}'] = "TOP 20 VARIANCES BY IMPACT"
    ws[f'A{row}'].font = Font(bold=True, size=12)

    row += 1
    headers = ["Employee ID", "Employee Name", "Wage Type", "Description", "Prior", "Current", "Variance $", "Variance %", "Risk"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for var in summary['top_variances']:
        row += 1
        ws.cell(row=row, column=1, value=var['emp_id'])
        ws.cell(row=row, column=2, value=var['emp_name'])
        ws.cell(row=row, column=3, value=var['wage_type'])
        ws.cell(row=row, column=4, value=var['wage_type_desc'])

        cell = ws.cell(row=row, column=5, value=var['prior_amount'])
        cell.number_format = '$#,##0.00'

        cell = ws.cell(row=row, column=6, value=var['current_amount'])
        cell.number_format = '$#,##0.00'

        cell = ws.cell(row=row, column=7, value=var['abs_variance'])
        cell.number_format = '$#,##0.00'

        cell = ws.cell(row=row, column=8, value=var['pct_variance'])
        cell.number_format = '0.0"%"'

        risk = var.get('risk_level', 'NONE')
        risk_cell = ws.cell(row=row, column=9, value=risk)
        if risk in RISK_FILLS:
            risk_cell.fill = RISK_FILLS[risk]

    # Adjust column widths
    for col_letter, width in [('A', 16), ('B', 18), ('C', 12), ('D', 25), ('E', 13), ('F', 13), ('G', 14), ('H', 12), ('I', 10)]:
        ws.column_dimensions[col_letter].width = width


def create_wage_type_sheet(output, wb):
    """Create By Wage Type sheet"""
    ws = wb.create_sheet("By Wage Type")

    headers = ["Wage Type", "Prior Total", "Current Total", "Abs Variance", "Pct Variance", "Flagged Items", "Total Items"]
    add_header_row(ws, headers, row=1)

    for row_idx, item in enumerate(output['by_wage_type'], 2):
        ws.cell(row=row_idx, column=1, value=item.get('wage_type', ''))

        cell = ws.cell(row=row_idx, column=2, value=item['prior_total'])
        cell.number_format = '$#,##0.00'

        cell = ws.cell(row=row_idx, column=3, value=item['current_total'])
        cell.number_format = '$#,##0.00'

        cell = ws.cell(row=row_idx, column=4, value=item['abs_variance'])
        cell.number_format = '$#,##0.00'

        cell = ws.cell(row=row_idx, column=5, value=item['pct_variance'])
        cell.number_format = '0.0"%"'

        ws.cell(row=row_idx, column=6, value=item['flagged_count'])
        ws.cell(row=row_idx, column=7, value=item['total_count'])

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 15


def create_cost_center_sheet(output, wb):
    """Create By Cost Center sheet"""
    ws = wb.create_sheet("By Cost Center")

    headers = ["Cost Center", "Prior Total", "Current Total", "Abs Variance", "Pct Variance", "Flagged Items", "Total Items"]
    add_header_row(ws, headers, row=1)

    for row_idx, item in enumerate(output['by_cost_center'], 2):
        ws.cell(row=row_idx, column=1, value=item.get('cost_center', ''))

        cell = ws.cell(row=row_idx, column=2, value=item['prior_total'])
        cell.number_format = '$#,##0.00'

        cell = ws.cell(row=row_idx, column=3, value=item['current_total'])
        cell.number_format = '$#,##0.00'

        cell = ws.cell(row=row_idx, column=4, value=item['abs_variance'])
        cell.number_format = '$#,##0.00'

        cell = ws.cell(row=row_idx, column=5, value=item['pct_variance'])
        cell.number_format = '0.0"%"'

        ws.cell(row=row_idx, column=6, value=item['flagged_count'])
        ws.cell(row=row_idx, column=7, value=item['total_count'])

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 15


def create_department_sheet(output, wb):
    """Create By Department sheet"""
    ws = wb.create_sheet("By Department")

    headers = ["Department", "Prior Total", "Current Total", "Abs Variance", "Pct Variance", "Flagged Items", "Total Items"]
    add_header_row(ws, headers, row=1)

    for row_idx, item in enumerate(output['by_department'], 2):
        ws.cell(row=row_idx, column=1, value=item.get('department', ''))

        cell = ws.cell(row=row_idx, column=2, value=item['prior_total'])
        cell.number_format = '$#,##0.00'

        cell = ws.cell(row=row_idx, column=3, value=item['current_total'])
        cell.number_format = '$#,##0.00'

        cell = ws.cell(row=row_idx, column=4, value=item['abs_variance'])
        cell.number_format = '$#,##0.00'

        cell = ws.cell(row=row_idx, column=5, value=item['pct_variance'])
        cell.number_format = '0.0"%"'

        ws.cell(row=row_idx, column=6, value=item['flagged_count'])
        ws.cell(row=row_idx, column=7, value=item['total_count'])

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 15


def create_detail_sheet(output, wb):
    """Create Detail sheet with all variance records including risk level"""
    ws = wb.create_sheet("Detail")

    headers = ["Employee ID", "Employee Name", "Cost Center", "Department", "Wage Type", "Description",
               "Prior Amount", "Current Amount", "Abs Variance", "Pct Variance", "Type", "Risk", "Flagged"]
    add_header_row(ws, headers, row=1)

    for row_idx, var in enumerate(output['all_variances'], 2):
        ws.cell(row=row_idx, column=1, value=var['emp_id'])
        ws.cell(row=row_idx, column=2, value=var['emp_name'])
        ws.cell(row=row_idx, column=3, value=var['cost_center'])
        ws.cell(row=row_idx, column=4, value=var['department'])
        ws.cell(row=row_idx, column=5, value=var['wage_type'])
        ws.cell(row=row_idx, column=6, value=var['wage_type_desc'])

        cell = ws.cell(row=row_idx, column=7, value=var['prior_amount'])
        cell.number_format = '$#,##0.00'

        cell = ws.cell(row=row_idx, column=8, value=var['current_amount'])
        cell.number_format = '$#,##0.00'

        cell = ws.cell(row=row_idx, column=9, value=var['abs_variance'])
        cell.number_format = '$#,##0.00'

        cell = ws.cell(row=row_idx, column=10, value=var['pct_variance'])
        cell.number_format = '0.0"%"'

        ws.cell(row=row_idx, column=11, value=var['variance_type'])

        risk = var.get('risk_level', 'NONE')
        risk_cell = ws.cell(row=row_idx, column=12, value=risk)
        if risk in RISK_FILLS:
            risk_cell.fill = RISK_FILLS[risk]

        ws.cell(row=row_idx, column=13, value="Yes" if var['flagged'] else "No")

        # Highlight flagged rows
        if var['flagged']:
            for col in range(1, 14):
                if col != 12:  # Don't override risk color
                    ws.cell(row=row_idx, column=col).fill = FLAG_FILL

    for col in range(1, 14):
        ws.column_dimensions[get_column_letter(col)].width = 14


def create_anomalies_sheet(output, wb):
    """Create dedicated Anomalies sheet with all detected anomalies, risk-colored and grouped"""
    ws = wb.create_sheet("Anomalies")

    anomalies = output.get('anomalies', [])

    # Title
    ws['A1'] = "ANOMALY DETECTION RESULTS"
    ws['A1'].font = Font(bold=True, size=14)

    total = len(anomalies)
    ws['A2'] = f"Total anomalies detected: {total}"
    ws['A2'].font = Font(bold=True)

    if total == 0:
        ws['A4'] = "No anomalies detected in this analysis."
        ws.column_dimensions['A'].width = 40
        return

    # ── Summary by type ─────────────────────────────────────────
    type_counts = defaultdict(int)
    for a in anomalies:
        type_counts[a['anomaly_type']] += 1

    row = 4
    ws[f'A{row}'] = "ANOMALY SUMMARY"
    ws[f'A{row}'].font = Font(bold=True, size=11)

    row += 1
    add_header_row(ws, ["Anomaly Type", "Count", "Highest Risk"], row=row)

    # Determine highest risk per type
    type_max_risk = {}
    risk_order = {'HIGH': 3, 'MEDIUM': 2, 'LOW': 1, 'NONE': 0}
    for a in anomalies:
        atype = a['anomaly_type']
        risk = a['risk_level']
        if atype not in type_max_risk or risk_order.get(risk, 0) > risk_order.get(type_max_risk[atype], 0):
            type_max_risk[atype] = risk

    for atype in ['NEW_HIRE', 'TERMINATION', 'GROSS_PAY_ANOMALY', 'WAGE_TYPE_APPEARED',
                   'WAGE_TYPE_DISAPPEARED', 'COST_CENTER_SHIFT', 'Z_SCORE_OUTLIER']:
        count = type_counts.get(atype, 0)
        if count > 0:
            row += 1
            label = ANOMALY_LABELS.get(atype, atype)
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=count)
            max_risk = type_max_risk.get(atype, 'NONE')
            risk_cell = ws.cell(row=row, column=3, value=max_risk)
            if max_risk in RISK_FILLS:
                risk_cell.fill = RISK_FILLS[max_risk]

    # ── Detail table ────────────────────────────────────────────
    row += 2
    ws[f'A{row}'] = "ANOMALY DETAILS"
    ws[f'A{row}'].font = Font(bold=True, size=11)

    row += 1
    detail_headers = ["Risk", "Type", "Employee ID", "Employee Name", "Cost Center", "Dept",
                      "Detail", "Impact $", "Investigation Notes"]
    add_header_row(ws, detail_headers, row=row)

    # Sort: HIGH first, then MEDIUM, then LOW
    sorted_anomalies = sorted(anomalies, key=lambda a: (-risk_order.get(a['risk_level'], 0), a['anomaly_type'], a['emp_id']))

    for a in sorted_anomalies:
        row += 1
        risk = a['risk_level']
        risk_cell = ws.cell(row=row, column=1, value=risk)
        if risk in RISK_FILLS:
            risk_cell.fill = RISK_FILLS[risk]

        ws.cell(row=row, column=2, value=ANOMALY_LABELS.get(a['anomaly_type'], a['anomaly_type']))
        ws.cell(row=row, column=3, value=a['emp_id'])
        ws.cell(row=row, column=4, value=a['emp_name'])
        ws.cell(row=row, column=5, value=a.get('cost_center', ''))
        ws.cell(row=row, column=6, value=a.get('department', ''))
        ws.cell(row=row, column=7, value=a['detail'])

        impact = a.get('impact', 0)
        cell = ws.cell(row=row, column=8, value=impact)
        cell.number_format = '$#,##0.00'

        ws.cell(row=row, column=9, value=a.get('investigation', ''))

    # Column widths
    col_widths = {'A': 8, 'B': 22, 'C': 14, 'D': 20, 'E': 14, 'F': 8, 'G': 55, 'H': 14, 'I': 60}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width


def main():
    parser = argparse.ArgumentParser(description='Generate variance report from analysis JSON')
    parser.add_argument('json_file', nargs='?', help='Input JSON file from analyze_variance.py')
    parser.add_argument('--output', type=str, default='variance_report.xlsx',
                        help='Output XLSX file path')

    args = parser.parse_args()

    if not args.json_file:
        parser.print_help()
        sys.exit(1)

    # Load JSON
    print(f"Loading analysis results: {args.json_file}")
    with open(args.json_file, 'r') as f:
        output = json.load(f)

    # Create workbook
    print("Creating report workbook...")
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create sheets
    print("  Creating Summary sheet...")
    create_summary_sheet(output, wb)

    print("  Creating By Wage Type sheet...")
    create_wage_type_sheet(output, wb)

    print("  Creating By Cost Center sheet...")
    create_cost_center_sheet(output, wb)

    print("  Creating By Department sheet...")
    create_department_sheet(output, wb)

    print("  Creating Detail sheet...")
    create_detail_sheet(output, wb)

    anomaly_count = len(output.get('anomalies', []))
    print(f"  Creating Anomalies sheet ({anomaly_count} anomalies)...")
    create_anomalies_sheet(output, wb)

    # Write workbook
    print(f"Writing report to {args.output}")
    wb.save(args.output)

    print("Report generation complete!")
    print(f"  Sheets: Summary, By Wage Type, By Cost Center, By Department, Detail, Anomalies")
    if anomaly_count > 0:
        print(f"  Anomalies: {anomaly_count} items flagged for investigation")

if __name__ == '__main__':
    main()
