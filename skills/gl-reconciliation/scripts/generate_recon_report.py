#!/usr/bin/env python3
"""
Generate Multi-Sheet Reconciliation Report

Reads JSON output from reconcile_payroll_gl.py and generates a comprehensive
multi-sheet Excel workbook with formatted summaries, detailed reconciliations,
and unmatched/reconciling items.

Usage:
    python generate_recon_report.py reconciliation_results.json --output recon_report.xlsx
"""

import argparse
import json
import sys
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers
)
from openpyxl.utils import get_column_letter


def format_currency(cell):
    """Format cell as currency."""
    cell.number_format = '$#,##0.00'


def format_percentage(cell):
    """Format cell as percentage."""
    cell.number_format = '0.0%'


def format_header(cell):
    """Format as header cell."""
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def format_subheader(cell):
    """Format as subheader cell."""
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")


def format_label(cell):
    """Format as label cell."""
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def format_matched(cell):
    """Format as matched (green)."""
    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    cell.font = Font(color="548235")


def format_unmatched(cell):
    """Format as unmatched (red)."""
    cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    cell.font = Font(color="C5504D")


def format_reconciling(cell):
    """Format as reconciling item (yellow)."""
    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    cell.font = Font(color="9C6500")


class ReconReportGenerator:
    def __init__(self, recon_data: Dict[str, Any]):
        self.data = recon_data
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate_summary_sheet(self):
        """Generate summary sheet with overview metrics."""
        ws = self.wb.create_sheet("Summary", 0)
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30

        row = 1
        ws[f'A{row}'] = "Payroll-to-GL Reconciliation Summary"
        format_header(ws[f'A{row}'])
        ws.merge_cells(f'A{row}:C{row}')
        row += 1

        ws[f'A{row}'] = "Generated"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row += 2

        # Summary metrics
        summary = self.data.get("reconciliation_summary", {})

        ws[f'A{row}'] = "Total Payroll Amount"
        format_label(ws[f'A{row}'])
        ws[f'B{row}'] = summary.get("total_payroll_amount", 0)
        format_currency(ws[f'B{row}'])
        row += 1

        ws[f'A{row}'] = "Total GL Amount"
        format_label(ws[f'A{row}'])
        ws[f'B{row}'] = summary.get("total_gl_amount", 0)
        format_currency(ws[f'B{row}'])
        row += 1

        ws[f'A{row}'] = "Total Variance"
        format_label(ws[f'A{row}'])
        ws[f'B{row}'] = summary.get("total_variance", 0)
        format_currency(ws[f'B{row}'])
        if summary.get("total_variance", 0) == 0:
            format_matched(ws[f'B{row}'])
        else:
            format_unmatched(ws[f'B{row}'])
        row += 2

        # Match statistics
        ws[f'A{row}'] = "Match Rate"
        format_label(ws[f'A{row}'])
        match_rate = summary.get("match_rate_percent", 0) / 100.0
        ws[f'B{row}'] = match_rate
        format_percentage(ws[f'B{row}'])
        if match_rate >= 0.95:
            format_matched(ws[f'B{row}'])
        elif match_rate >= 0.85:
            format_reconciling(ws[f'B{row}'])
        else:
            format_unmatched(ws[f'B{row}'])
        row += 1

        ws[f'A{row}'] = "Matched Items"
        format_label(ws[f'A{row}'])
        ws[f'B{row}'] = summary.get("matched_count", 0)
        format_matched(ws[f'B{row}'])
        row += 1

        ws[f'A{row}'] = "Unmatched Items"
        format_label(ws[f'A{row}'])
        ws[f'B{row}'] = summary.get("unmatched_count", 0)
        if summary.get("unmatched_count", 0) > 0:
            format_unmatched(ws[f'B{row}'])
        row += 1

        ws[f'A{row}'] = "Reconciling Items"
        format_label(ws[f'A{row}'])
        ws[f'B{row}'] = summary.get("reconciling_items_count", 0)
        if summary.get("reconciling_items_count", 0) > 0:
            format_reconciling(ws[f'B{row}'])
        row += 2

        # Reconciliation type summaries
        ws[f'A{row}'] = "By Reconciliation Type"
        format_subheader(ws[f'A{row}'])
        ws.merge_cells(f'A{row}:C{row}')
        row += 1

        results = self.data.get("reconciliation_results", {})
        for recon_type, recon_data in results.items():
            ws[f'A{row}'] = recon_type.replace("_", " ").title()
            format_label(ws[f'A{row}'])
            row += 1

            ws[f'A{row}'] = "  Matched"
            ws[f'B{row}'] = len(recon_data.get("matched", []))
            format_matched(ws[f'B{row}'])
            row += 1

            ws[f'A{row}'] = "  Unmatched"
            ws[f'B{row}'] = len(recon_data.get("unmatched", []))
            if len(recon_data.get("unmatched", [])) > 0:
                format_unmatched(ws[f'B{row}'])
            row += 1

    def generate_gross_to_net_sheet(self):
        """Generate gross-to-net reconciliation sheet."""
        recon_data = self.data.get("reconciliation_results", {}).get("gross_to_net", {})
        if not recon_data:
            return

        ws = self.wb.create_sheet("Gross-to-Net")
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

        row = 1
        ws[f'A{row}'] = "Gross-to-Net Reconciliation"
        format_header(ws[f'A{row}'])
        ws.merge_cells(f'A{row}:D{row}')
        row += 2

        # Walkdown summary
        walkdown = recon_data.get("walkdown", {})

        ws[f'A{row}'] = "Payroll Walkdown"
        format_subheader(ws[f'A{row}'])
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        ws[f'A{row}'] = "Total Gross Salary"
        ws[f'B{row}'] = walkdown.get("total_gross", 0)
        format_currency(ws[f'B{row}'])
        row += 1

        ws[f'A{row}'] = "Total Deductions"
        ws[f'B{row}'] = walkdown.get("total_deductions", 0)
        format_currency(ws[f'B{row}'])
        row += 1

        ws[f'A{row}'] = "Net Pay (Payroll)"
        ws[f'B{row}'] = walkdown.get("total_net", 0)
        format_currency(ws[f'B{row}'])
        format_matched(ws[f'B{row}'])
        row += 2

        # GL comparison
        ws[f'A{row}'] = "GL Comparison"
        format_subheader(ws[f'A{row}'])
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        ws[f'A{row}'] = "Description"
        ws[f'B{row}'] = "Payroll Amount"
        ws[f'C{row}'] = "GL Amount"
        ws[f'D{row}'] = "Variance"
        for col in ['A', 'B', 'C', 'D']:
            format_subheader(ws[f'{col}{row}'])
        row += 1

        for item in recon_data.get("matched", []):
            if item.get("type") == "net_pay":
                ws[f'A{row}'] = "Net Pay"
                ws[f'B{row}'] = item.get("payroll_amount", 0)
                ws[f'C{row}'] = item.get("gl_amount", 0)
                ws[f'D{row}'] = item.get("variance", 0)
                for col in ['B', 'C', 'D']:
                    format_currency(ws[f'{col}{row}'])
                    format_matched(ws[f'{col}{row}'])
                row += 1

        for item in recon_data.get("matched", []):
            if item.get("wage_type"):
                ws[f'A{row}'] = item.get("description", item.get("wage_type"))
                ws[f'B{row}'] = item.get("payroll_amount", 0)
                ws[f'C{row}'] = item.get("gl_amount", 0)
                ws[f'D{row}'] = item.get("variance", 0)
                for col in ['B', 'C', 'D']:
                    format_currency(ws[f'{col}{row}'])
                    format_matched(ws[f'{col}{row}'])
                row += 1

        for item in recon_data.get("unmatched", []):
            if item.get("wage_type"):
                ws[f'A{row}'] = item.get("description", item.get("wage_type"))
                ws[f'B{row}'] = item.get("payroll_amount", 0)
                ws[f'C{row}'] = item.get("gl_amount", 0)
                ws[f'D{row}'] = item.get("variance", 0)
                for col in ['B', 'C', 'D']:
                    format_currency(ws[f'{col}{row}'])
                    format_unmatched(ws[f'{col}{row}'])
                row += 1

    def generate_employer_costs_sheet(self):
        """Generate employer costs reconciliation sheet."""
        recon_data = self.data.get("reconciliation_results", {}).get("employer_costs", {})
        if not recon_data:
            return

        ws = self.wb.create_sheet("Employer Costs")
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

        row = 1
        ws[f'A{row}'] = "Employer Cost Reconciliation"
        format_header(ws[f'A{row}'])
        ws.merge_cells(f'A{row}:D{row}')
        row += 2

        # Header
        ws[f'A{row}'] = "Cost Type"
        ws[f'B{row}'] = "Payroll Amount"
        ws[f'C{row}'] = "GL Amount"
        ws[f'D{row}'] = "Variance"
        for col in ['A', 'B', 'C', 'D']:
            format_subheader(ws[f'{col}{row}'])
        row += 1

        by_type = recon_data.get("by_cost_type", {})
        for cost_type, data in by_type.items():
            ws[f'A{row}'] = cost_type
            ws[f'B{row}'] = data.get("payroll_amount", 0)
            ws[f'C{row}'] = data.get("gl_amount", 0)
            ws[f'D{row}'] = data.get("variance", 0)

            for col in ['B', 'C', 'D']:
                format_currency(ws[f'{col}{row}'])

            if data.get("matched"):
                for col in ['B', 'C', 'D']:
                    format_matched(ws[f'{col}{row}'])
            else:
                for col in ['B', 'C', 'D']:
                    format_unmatched(ws[f'{col}{row}'])

            row += 1

    def generate_tax_liabilities_sheet(self):
        """Generate tax liabilities reconciliation sheet."""
        recon_data = self.data.get("reconciliation_results", {}).get("tax_liabilities", {})
        if not recon_data:
            return

        ws = self.wb.create_sheet("Tax Liabilities")
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

        row = 1
        ws[f'A{row}'] = "Tax Liability Reconciliation"
        format_header(ws[f'A{row}'])
        ws.merge_cells(f'A{row}:D{row}')
        row += 2

        # Header
        ws[f'A{row}'] = "Jurisdiction"
        ws[f'B{row}'] = "Payroll Amount"
        ws[f'C{row}'] = "GL Amount"
        ws[f'D{row}'] = "Variance"
        for col in ['A', 'B', 'C', 'D']:
            format_subheader(ws[f'{col}{row}'])
        row += 1

        by_juris = recon_data.get("by_jurisdiction", {})
        for jurisdiction, data in by_juris.items():
            ws[f'A{row}'] = jurisdiction
            ws[f'B{row}'] = data.get("payroll_amount", 0)
            ws[f'C{row}'] = data.get("gl_amount", 0)
            ws[f'D{row}'] = data.get("variance", 0)

            for col in ['B', 'C', 'D']:
                format_currency(ws[f'{col}{row}'])

            if data.get("matched"):
                for col in ['B', 'C', 'D']:
                    format_matched(ws[f'{col}{row}'])
            else:
                for col in ['B', 'C', 'D']:
                    format_unmatched(ws[f'{col}{row}'])

            row += 1

    def generate_cost_center_sheet(self):
        """Generate cost center allocation sheet."""
        recon_data = self.data.get("reconciliation_results", {}).get("cost_center_allocation", {})
        if not recon_data:
            return

        ws = self.wb.create_sheet("Cost Center")
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

        row = 1
        ws[f'A{row}'] = "Cost Center Allocation Reconciliation"
        format_header(ws[f'A{row}'])
        ws.merge_cells(f'A{row}:D{row}')
        row += 2

        # Header
        ws[f'A{row}'] = "Cost Center"
        ws[f'B{row}'] = "Payroll Amount"
        ws[f'C{row}'] = "GL Amount"
        ws[f'D{row}'] = "Variance"
        for col in ['A', 'B', 'C', 'D']:
            format_subheader(ws[f'{col}{row}'])
        row += 1

        by_cc = recon_data.get("by_cost_center", {})
        for cost_center in sorted(by_cc.keys()):
            data = by_cc[cost_center]
            ws[f'A{row}'] = cost_center
            ws[f'B{row}'] = data.get("payroll_amount", 0)
            ws[f'C{row}'] = data.get("gl_amount", 0)
            ws[f'D{row}'] = data.get("variance", 0)

            for col in ['B', 'C', 'D']:
                format_currency(ws[f'{col}{row}'])

            if data.get("matched"):
                for col in ['B', 'C', 'D']:
                    format_matched(ws[f'{col}{row}'])
            else:
                for col in ['B', 'C', 'D']:
                    format_unmatched(ws[f'{col}{row}'])

            row += 1

    def generate_unmatched_sheet(self):
        """Generate unmatched items sheet."""
        ws = self.wb.create_sheet("Unmatched Items")
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 50

        row = 1
        ws[f'A{row}'] = "Unmatched Items"
        format_header(ws[f'A{row}'])
        ws.merge_cells(f'A{row}:D{row}')
        row += 2

        # Header
        ws[f'A{row}'] = "Recon Type"
        ws[f'B{row}'] = "Description"
        ws[f'C{row}'] = "Amount"
        ws[f'D{row}'] = "Reason"
        for col in ['A', 'B', 'C', 'D']:
            format_subheader(ws[f'{col}{row}'])
        row += 1

        results = self.data.get("reconciliation_results", {})
        for recon_type, recon_data in results.items():
            for item in recon_data.get("unmatched", []):
                ws[f'A{row}'] = recon_type.replace("_", " ").title()

                desc = item.get("description") or item.get("wage_type") or item.get("cost_center") or "Unknown"
                ws[f'B{row}'] = desc

                amount = item.get("payroll_amount") or item.get("gl_amount") or item.get("variance") or 0
                ws[f'C{row}'] = amount
                format_currency(ws[f'C{row}'])

                ws[f'D{row}'] = item.get("reason", "Variance exceeds tolerance")

                format_unmatched(ws[f'A{row}'])
                format_unmatched(ws[f'B{row}'])
                format_unmatched(ws[f'C{row}'])
                format_unmatched(ws[f'D{row}'])

                row += 1

    def generate_reconciling_items_sheet(self):
        """Generate reconciling items sheet."""
        ws = self.wb.create_sheet("Reconciling Items")
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 50

        row = 1
        ws[f'A{row}'] = "Reconciling Items"
        format_header(ws[f'A{row}'])
        ws.merge_cells(f'A{row}:C{row}')
        row += 2

        # Header
        ws[f'A{row}'] = "Type"
        ws[f'B{row}'] = "Description"
        ws[f'C{row}'] = "Resolution Steps"
        for col in ['A', 'B', 'C']:
            format_subheader(ws[f'{col}{row}'])
        row += 1

        reconciling_items = self.data.get("reconciling_items", [])
        for item in reconciling_items:
            ws[f'A{row}'] = item.get("type", "Unknown").title()
            ws[f'B{row}'] = item.get("description", "")

            resolution_steps = item.get("resolution_steps", [])
            resolution_text = "\n".join(resolution_steps)
            ws[f'C{row}'] = resolution_text

            ws[f'C{row}'].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = len(resolution_steps) * 15

            for col in ['A', 'B', 'C']:
                format_reconciling(ws[f'{col}{row}'])

            row += 1

    def generate(self) -> Path:
        """Generate complete workbook."""
        self.generate_summary_sheet()
        self.generate_gross_to_net_sheet()
        self.generate_employer_costs_sheet()
        self.generate_tax_liabilities_sheet()
        self.generate_cost_center_sheet()
        self.generate_unmatched_sheet()
        self.generate_reconciling_items_sheet()

        return self.wb


def main():
    parser = argparse.ArgumentParser(
        description="Generate multi-sheet reconciliation report from JSON results"
    )
    parser.add_argument("json_file", help="Path to JSON reconciliation results file")
    parser.add_argument(
        "--output",
        default="reconciliation_report.xlsx",
        help="Output Excel file (default: reconciliation_report.xlsx)"
    )

    args = parser.parse_args()

    # Load JSON
    try:
        with open(args.json_file, 'r') as f:
            data = json.load(f)
    except Exception as e:
        print(f"Error loading JSON file: {e}", file=sys.stderr)
        sys.exit(1)

    # Generate report
    generator = ReconReportGenerator(data)
    wb = generator.generate()

    # Save
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Report generated successfully: {output_path}")


if __name__ == "__main__":
    main()
